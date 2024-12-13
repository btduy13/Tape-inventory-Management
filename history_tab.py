import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tab_base import TabBase
import pandas as pd
from datetime import datetime, timedelta
import os
from tkcalendar import DateEntry
from database import BangKeoInOrder, TrucInOrder
import logging
from sqlalchemy import func
from excel_import import export_template, import_data
from openpyxl import Workbook, load_workbook

class HistoryTab(TabBase):
    def __init__(self, notebook, parent_form):
        super().__init__(parent_form)
        self.tab = ttk.Frame(notebook)
        notebook.add(self.tab, text="Lịch sử")
        self.db_session = parent_form.db_session
        
        # Define standard date formats
        self.DATE_FORMAT = '%d/%m/%Y'
        self.DATETIME_FORMAT = '%d/%m/%Y %H:%M:%S'
        
        # Create main frame with padding
        main_frame = ttk.Frame(self.tab, padding="15 15 15 15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Initialize data storage
        self.bang_keo_in_data = []
        self.truc_in_data = []
        
        # Add these lines
        self.all_bang_keo_in_items = []  # Store all bang keo items
        self.all_truc_in_items = []   # Store all truc in items
        
        # Create edit frame (initially hidden)
        self.edit_frame = ttk.LabelFrame(main_frame, text="Chỉnh sửa thông tin", padding="10 5 10 5")
        self.edit_entries = {}
        self.current_edit_item = None
        self.current_edit_type = None
        
        # Build the UI components
        self.build_ui(main_frame)
        
        # Load data from database
        self.load_data_from_db()

    def parse_date_string(self, date_string):
        """Parse a date string to datetime object"""
        try:
            if ' ' in date_string:
                return datetime.strptime(date_string, self.DATETIME_FORMAT)
            return datetime.strptime(date_string, self.DATE_FORMAT)
        except ValueError as e:
            logging.error(f"Date parsing error: {str(e)}")
            return None

    def load_data_from_db(self):
        """Load data from database"""
        try:
            # Load bang keo orders
            bang_keo_in_orders = self.db_session.query(BangKeoInOrder).all()
            for order in bang_keo_in_orders:
                order_data = {
                    'id': order.id,
                    'thoi_gian': order.thoi_gian.strftime(self.DATE_FORMAT),
                    'ten_hang': order.ten_hang,
                    'ngay_du_kien': order.ngay_du_kien.strftime(self.DATE_FORMAT),
                    'quy_cach_mm': order.quy_cach_mm,
                    'quy_cach_m': order.quy_cach_m,
                    'quy_cach_mic': order.quy_cach_mic,
                    'cuon_cay': order.cuon_cay,
                    'so_luong': order.so_luong,
                    'phi_sl': order.phi_sl,
                    'mau_keo': order.mau_keo,
                    'phi_keo': order.phi_keo,
                    'mau_sac': order.mau_sac,
                    'phi_mau': order.phi_mau,
                    'phi_size': order.phi_size,
                    'phi_cat': order.phi_cat,
                    'don_gia_von': order.don_gia_von,
                    'don_gia_goc': order.don_gia_goc,
                    'thanh_tien_goc': order.thanh_tien_goc,
                    'don_gia_ban': order.don_gia_ban,
                    'thanh_tien_ban': order.thanh_tien_ban,
                    'tien_coc': order.tien_coc,
                    'cong_no_khach': order.cong_no_khach,
                    'ctv': order.ctv,
                    'hoa_hong': order.hoa_hong,
                    'tien_hoa_hong': order.tien_hoa_hong,
                    'loi_giay': order.loi_giay,
                    'thung_bao': order.thung_bao,
                    'loi_nhuan': order.loi_nhuan,
                    'da_giao': order.da_giao,
                    'da_tat_toan': order.da_tat_toan
                }
                self.add_order('Băng keo in', order_data)
            
            # Load truc in orders
            truc_in_orders = self.db_session.query(TrucInOrder).all()
            for order in truc_in_orders:
                order_data = {
                    'id': order.id,
                    'thoi_gian': order.thoi_gian.strftime(self.DATE_FORMAT),
                    'ten_hang': order.ten_hang,
                    'ngay_du_kien': order.ngay_du_kien.strftime(self.DATE_FORMAT),
                    'quy_cach': order.quy_cach,
                    'so_luong': order.so_luong,
                    'mau_sac': order.mau_sac,
                    'mau_keo': order.mau_keo,
                    'don_gia_goc': order.don_gia_goc,
                    'thanh_tien': order.thanh_tien,
                    'don_gia_ban': order.don_gia_ban,
                    'thanh_tien_ban': order.thanh_tien_ban,
                    'cong_no_khach': order.cong_no_khach,
                    'ctv': order.ctv,
                    'hoa_hong': order.hoa_hong,
                    'tien_hoa_hong': order.tien_hoa_hong,
                    'loi_nhuan': order.loi_nhuan,
                    'da_giao': order.da_giao,
                    'da_tat_toan': order.da_tat_toan
                }
                self.add_order('Trục in', order_data)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi tải dữ liệu: {str(e)}")
            self.update_status("Lỗi khi tải dữ liệu từ database")

    def delete_selected(self):
        """Delete selected items from treeview and database"""
        try:
            current_tab = self.category_notebook.select()
            if current_tab == str(self.bang_keo_in_frame):
                tree = self.bang_keo_in_tree
                model = BangKeoInOrder
            else:
                tree = self.truc_in_tree
                model = TrucInOrder

            selected_items = tree.selection()
            if not selected_items:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn ít nhất một dòng để xóa")
                return

            if messagebox.askyesno("Xác nhận", "Bạn có chắc muốn xóa các dòng đã chọn?"):
                for item in selected_items:
                    try:
                        # Get the ID from the first column of the tree item
                        item_values = tree.item(item)['values']
                        record_id = item_values[0]  # ID is now the first column
                        
                        # Delete from database
                        record = self.db_session.query(model).filter_by(id=record_id).first()
                        if record:
                            self.db_session.delete(record)
                            self.db_session.commit()
                            logging.info(f"Successfully deleted record {record_id} from database")
                        
                        # Delete from tree
                        tree.delete(item)
                        logging.info(f"Successfully deleted item from tree")
                    except Exception as e:
                        logging.error(f"Error deleting item: {str(e)}")
                        self.db_session.rollback()
                        continue

                try:
                    self.update_status("Đã xóa các dòng đã chọn")
                    
                    # Update the stored items lists
                    if current_tab == str(self.bang_keo_in_frame):
                        self.all_bang_keo_in_items = [
                            (item, self.bang_keo_in_tree.item(item)['values'])
                            for item in self.bang_keo_in_tree.get_children()
                        ]
                    else:
                        self.all_truc_in_items = [
                            (item, self.truc_in_tree.item(item)['values'])
                            for item in self.truc_in_tree.get_children()
                        ]
                        
                    # Update other tabs if they exist
                    if hasattr(self.parent_form, 'thong_ke_tab'):
                        self.parent_form.thong_ke_tab.load_data()
                        
                except Exception as e:
                    logging.error(f"Error updating UI after deletion: {str(e)}")
                    raise

        except Exception as e:
            self.db_session.rollback()
            logging.error(f"Error in delete_selected: {str(e)}")
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi xóa: {str(e)}")
            self.update_status("Lỗi khi xóa dữ liệu")

    def build_ui(self, main_frame):
        # Title with better styling
        title_frame = ttk.Frame(main_frame)
        title_frame.pack(fill=tk.X, pady=(0, 10))
        title_label = ttk.Label(title_frame, text="LỊCH SỬ ĐƠN HÀNG", font=('Arial', 16, 'bold'))
        title_label.pack(pady=10)
        
        # Search and filter frame
        search_filter_frame = ttk.LabelFrame(main_frame, text="Tìm kiếm và lọc", padding="5 5 5 5")
        search_filter_frame.pack(fill=tk.X, padx=5, pady=(0, 10))
        
        # Date filter frame
        date_frame = ttk.Frame(search_filter_frame)
        date_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # From date
        ttk.Label(date_frame, text="Từ ngày:").pack(side=tk.LEFT, padx=5)
        self.from_date = DateEntry(date_frame, width=12, background='darkblue',
                                 foreground='white', borderwidth=2,
                                 date_pattern='dd-mm-yyyy', locale='vi_VN')
        self.from_date.pack(side=tk.LEFT, padx=5)
        
        # To date
        ttk.Label(date_frame, text="Đến ngày:").pack(side=tk.LEFT, padx=5)
        self.to_date = DateEntry(date_frame, width=12, background='darkblue',
                               foreground='white', borderwidth=2,
                               date_pattern='dd-mm-yyyy', locale='vi_VN')
        self.to_date.pack(side=tk.LEFT, padx=5)
        
        # Filter button
        filter_btn = ttk.Button(date_frame, text="Lọc", command=self.apply_filters)
        filter_btn.pack(side=tk.LEFT, padx=5)
        
        # Reset filter button
        reset_btn = ttk.Button(date_frame, text="Đặt lại", command=self.reset_filters)
        reset_btn.pack(side=tk.LEFT, padx=5)
        
        # Search frame
        search_frame = ttk.Frame(search_filter_frame)
        search_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Search entry
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.on_search)
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=40)
        search_entry.pack(side=tk.LEFT, padx=5)
        ttk.Label(search_frame, text="(Tìm theo tên hàng, thời gian, CTV)").pack(side=tk.LEFT, padx=5)
        
        # Instructions frame
        instruction_frame = ttk.LabelFrame(main_frame, text="Hướng dẫn sử dụng", padding="10 5 10 5")
        instruction_frame.pack(fill=tk.X, padx=5, pady=5)
        
        instructions = [
            "- Click chuột để chọn một dòng",
            "- Giữ Ctrl + Click để chọn nhiều dòng riêng lẻ",
            "- Ctrl + A để chọn tất cả các dòng",
            "- Double-click để chỉnh sửa thông tin",
        ]
        
        for instruction in instructions:
            ttk.Label(instruction_frame, text=instruction).pack(anchor=tk.W, padx=5, pady=2)
        
        # Create notebook for categories
        self.category_notebook = ttk.Notebook(main_frame)
        self.category_notebook.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Create frames for each category
        self.bang_keo_in_frame = ttk.Frame(self.category_notebook, padding="5 5 5 5")
        self.truc_in_frame = ttk.Frame(self.category_notebook, padding="5 5 5 5")
        
        self.category_notebook.add(self.bang_keo_in_frame, text="Băng keo in")
        self.category_notebook.add(self.truc_in_frame, text="Trục in")
        
        # Create treeviews
        self.create_bang_keo_in_tree()
        self.create_truc_in_tree()
        
        # Create button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, padx=5, pady=10, side=tk.BOTTOM)
        
        # Configure grid to make buttons stick during resizing
        button_frame.columnconfigure((0, 1, 2, 3), weight=1)
        
        # Add buttons with tooltips
        export_excel_btn = ttk.Button(button_frame, text="Xuất Excel", command=self.export_selected_to_excel, width=15)
        export_excel_btn.grid(row=0, column=0, padx=5, sticky='ew')
        self.create_tooltip(export_excel_btn, "Xuất dữ liệu đã chọn ra file Excel (Ctrl+E)")
        
        export_email_btn = ttk.Button(button_frame, text="Xuất Email", command=self.export_selected_to_email, width=15)
        export_email_btn.grid(row=0, column=1, padx=5, sticky='ew')
        self.create_tooltip(export_email_btn, "Xuất dữ liệu đã chọn thành email (Ctrl+M)")
        
        delete_btn = ttk.Button(button_frame, text="Xóa", command=self.delete_selected, width=15)
        delete_btn.grid(row=0, column=2, padx=5, sticky='ew')
        self.create_tooltip(delete_btn, "Xóa các dòng đã chọn (Delete)")
        
        refresh_btn = ttk.Button(button_frame, text="Làm mới", command=self.refresh_data, width=15)
        refresh_btn.grid(row=0, column=3, padx=5, sticky='ew')
        self.create_tooltip(refresh_btn, "Làm mới dữ liệu (F5)")
        
        # Status bar
        self.status_bar = ttk.Label(main_frame, text="", relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(fill=tk.X, side=tk.BOTTOM, pady=(5, 0))
        
        # Bind keyboard shortcuts
        self.bind_shortcuts()
        
        # Bind notebook tab change event
        self.category_notebook.bind('<<NotebookTabChanged>>', self.on_tab_changed)
        
    def create_tooltip(self, widget, text):
        """Create a tooltip for a widget"""
        def show_tooltip(event):
            tooltip = tk.Toplevel()
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
            
            label = ttk.Label(tooltip, text=text, background="#ffffe0", relief=tk.SOLID, borderwidth=1)
            label.pack()
            
            def hide_tooltip():
                tooltip.destroy()
            
            widget.tooltip = tooltip
            widget.bind('<Leave>', lambda e: hide_tooltip())
            tooltip.bind('<Leave>', lambda e: hide_tooltip())
        
        widget.bind('<Enter>', show_tooltip)

    def on_search(self, *args):
        """Filter the treeview based on search text and date range"""
        search_text = self.search_var.get().lower()
        current_tab = self.category_notebook.select()
        
        if current_tab == str(self.bang_keo_in_frame):
            tree = self.bang_keo_in_tree
            all_items = self.all_bang_keo_in_items
        else:
            tree = self.truc_in_tree
            all_items = self.all_truc_in_items

        # Clear current tree
        for item in tree.get_children():
            tree.delete(item)
        
        # Get date range if set
        from_date = self.from_date.get_date()
        to_date = self.to_date.get_date()

        # Reinsert matching items
        for _, values in all_items:
            try:
                # Check date range if dates are set
                date_match = True
                if from_date and to_date:
                    order_date = datetime.strptime(values[1], self.DATE_FORMAT).date()
                    expected_date = datetime.strptime(values[3], self.DATE_FORMAT).date()
                    date_match = (from_date <= order_date <= to_date or 
                                from_date <= expected_date <= to_date)

                # Check if any value contains search text
                text_match = True
                if search_text:
                    text_match = False
                    for value in values:
                        if str(value).lower().find(search_text) != -1:
                            text_match = True
                            break

                # Insert if both conditions are met
                if date_match and text_match:
                    tree.insert('', 'end', values=values)

            except (ValueError, IndexError) as e:
                logging.error(f"Error processing item in on_search: {str(e)}")
                continue

        # Apply alternating row colors
        self._apply_row_colors(tree)

    def _apply_row_colors(self, tree):
        """Apply alternating row colors to tree"""
        items = tree.get_children()
        for i, item in enumerate(items):
            if i % 2 == 0:
                tree.tag_configure('evenrow', background='#FFFFFF')
                tree.item(item, tags=('evenrow',))
            else:
                tree.tag_configure('oddrow', background='#F0F0F0')
                tree.item(item, tags=('oddrow',))

    def refresh_data(self):
        """Refresh data from database"""
        try:
            # Clear existing data
            self.bang_keo_in_tree.delete(*self.bang_keo_in_tree.get_children())
            self.truc_in_tree.delete(*self.truc_in_tree.get_children())
            self.bang_keo_in_data.clear()
            self.truc_in_data.clear()
            self.all_bang_keo_in_items.clear()
            self.all_truc_in_items.clear()
            
            # Reload data from database
            self.load_data_from_db()
            
            # Update status
            self.update_status("Đã cập nhật dữ liệu")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi cập nhật dữ liệu: {str(e)}")
            self.update_status("Lỗi khi cập nhật dữ liệu")

    def bind_shortcuts(self):
        """Bind keyboard shortcuts"""
        self.root.bind('<Control-f>', lambda e: self.focus_search())
        self.root.bind('<F5>', lambda e: self.refresh_data())
        self.root.bind('<Delete>', lambda e: self.delete_selected())
        self.root.bind('<Control-e>', lambda e: self.export_selected_to_excel())
        self.root.bind('<Control-m>', lambda e: self.export_selected_to_email())

    def focus_search(self):
        """Focus the search entry"""
        self.search_var.set('')  # Clear existing search
        for child in self.search_frame.winfo_children():
            if isinstance(child, ttk.Entry):
                child.focus_set()
                break

    def update_status(self, message):
        """Update the status bar with a message"""
        self.status_bar.config(text=message)
        self.root.after(3000, lambda: self.status_bar.config(text=""))  # Clear after 3 seconds

    def create_bang_keo_in_tree(self):
        columns = ('id', 'thoi_gian', 'ten_hang', 'ngay_du_kien', 'quy_cach_mm', 'quy_cach_m', 'quy_cach_mic', 
                  'cuon_cay', 'so_luong', 'phi_sl', 'mau_keo', 'phi_keo', 'mau_sac', 
                  'phi_mau', 'phi_size', 'phi_cat', 'don_gia_von', 'don_gia_goc', 
                  'thanh_tien_goc', 'don_gia_ban', 'thanh_tien_ban', 'tien_coc', 
                  'cong_no_khach', 'ctv', 'hoa_hong', 'tien_hoa_hong',
                  'loi_giay', 'thung_bao', 'loi_nhuan', 'da_giao', 'da_tat_toan')
        
        # Create container frame
        container = ttk.Frame(self.bang_keo_in_frame)
        container.pack(fill=tk.BOTH, expand=True)
        
        # Configure grid weights
        container.grid_columnconfigure(0, weight=1)
        container.grid_rowconfigure(0, weight=1)
        
        # Create treeview with alternating row colors
        style = ttk.Style()
        style.configure("Custom.Treeview",
                       background="#ffffff",
                       foreground="black",
                       fieldbackground="#ffffff",
                       rowheight=25)
        style.map("Custom.Treeview",
                 background=[("selected", "#0078D7")],
                 foreground=[("selected", "#ffffff")])
        
        self.bang_keo_in_tree = ttk.Treeview(container, columns=columns, show='headings',
                                         selectmode='extended', style="Custom.Treeview")
        
        # Define headings and column widths
        headings = {
            'id': 'ID',
            'thoi_gian': 'Thời gian', 'ten_hang': 'Tên hàng', 
            'ngay_du_kien': 'Ngày dự kiến giao',
            'quy_cach_mm': 'Quy cách (mm)', 'quy_cach_m': 'Quy cách (m)', 
            'quy_cach_mic': 'Quy cách (mic)', 'cuon_cay': 'Cuộn/cây',
            'so_luong': 'Số lượng', 'phi_sl': 'Phí SL', 'mau_keo': 'Màu keo',
            'phi_keo': 'Phí keo', 'mau_sac': 'Màu sắc', 'phi_mau': 'Phí màu',
            'phi_size': 'Phí size', 'phi_cat': 'Phí cắt',
            'don_gia_von': 'Đơn giá vốn', 'don_gia_goc': 'Đơn giá gốc',
            'thanh_tien_goc': 'Thành tiền gốc', 'don_gia_ban': 'Đơn giá bán',
            'thanh_tien_ban': 'Thành tiền bán', 'tien_coc': 'Tiền cọc',
            'cong_no_khach': 'Công nợ khách', 'ctv': 'CTV',
            'hoa_hong': 'Hoa hồng', 'tien_hoa_hong': 'Tiền hoa hồng',
            'loi_giay': 'Li giấy', 'thung_bao': 'Thùng/Bao',
            'loi_nhuan': 'Lợi nhuận',
            'da_giao': 'Đã giao',
            'da_tat_toan': 'Đã tất toán'
        }
        
        for col in columns:
            self.bang_keo_in_tree.heading(col, text=headings[col],
                                     command=lambda c=col: self.sort_treeview(self.bang_keo_in_tree, c, False))
            # Hide ID column
            if col == 'id':
                self.bang_keo_in_tree.column(col, width=50, stretch=False)
            else:
                # Adjust column widths based on content
                width = max(len(headings[col])*10, 100)
                self.bang_keo_in_tree.column(col, width=width, stretch=False)
        
        # Add scrollbars with modern styling
        y_scrollbar = ttk.Scrollbar(container, orient=tk.VERTICAL, command=self.bang_keo_in_tree.yview)
        x_scrollbar = ttk.Scrollbar(container, orient=tk.HORIZONTAL, command=self.bang_keo_in_tree.xview)
        self.bang_keo_in_tree.configure(yscrollcommand=y_scrollbar.set, xscrollcommand=x_scrollbar.set)
        
        # Grid layout
        self.bang_keo_in_tree.grid(row=0, column=0, sticky='nsew')
        y_scrollbar.grid(row=0, column=1, sticky='ns')
        x_scrollbar.grid(row=1, column=0, sticky='ew')
        
        # Bind events
        self.bang_keo_in_tree.bind('<Double-1>', lambda e: self.root.after(100, lambda: self.show_edit_form('bang_keo_in')))
        self.bang_keo_in_tree.bind('<Control-a>', lambda e: self.select_all(self.bang_keo_in_tree))
        
        # Add alternating row colors
        self.bang_keo_in_tree.tag_configure('oddrow', background='#F0F0F0')
        self.bang_keo_in_tree.tag_configure('evenrow', background='#FFFFFF')

    def sort_treeview(self, tree, col, reverse):
        """Sort treeview content when clicking on headers"""
        items = [(tree.set(item, col), item) for item in tree.get_children('')]
        
        # Convert to appropriate type for sorting
        def convert_value(value):
            try:
                # Try to convert to float first (for numeric values)
                return float(value.replace(',', ''))
            except:
                # If not possible, return as string
                return value.lower()
        
        items.sort(key=lambda x: convert_value(x[0]), reverse=reverse)
        
        # Rearrange items in sorted positions
        for index, (val, item) in enumerate(items):
            tree.move(item, '', index)
            # Update row colors
            tree.tag_configure('oddrow', background='#F0F0F0')
            tree.tag_configure('evenrow', background='#FFFFFF')
            if index % 2 == 0:
                tree.item(item, tags=('evenrow',))
            else:
                tree.item(item, tags=('oddrow',))
        
        # Reverse sort next time
        tree.heading(col, command=lambda: self.sort_treeview(tree, col, not reverse))
        
    def create_truc_in_tree(self):
        columns = ('id', 'thoi_gian', 'ten_hang', 'ngay_du_kien', 'quy_cach', 'so_luong', 'mau_sac',
                  'mau_keo', 'don_gia_goc', 'thanh_tien', 'don_gia_ban',
                  'thanh_tien_ban', 'cong_no_khach', 'ctv', 'hoa_hong',
                  'tien_hoa_hong', 'loi_nhuan', 'da_giao', 'da_tat_toan')
        
        # Create container frame for better performance
        container = ttk.Frame(self.truc_in_frame)
        container.pack(fill=tk.BOTH, expand=True)
        
        # Configure grid weights
        container.grid_columnconfigure(0, weight=1)
        container.grid_rowconfigure(0, weight=1)
        
        # Change selectmode to 'extended' to allow multiple selections
        self.truc_in_tree = ttk.Treeview(container, columns=columns, show='headings', selectmode='extended')
        
        # Define headings
        headings = {
            'id': 'ID',
            'thoi_gian': 'Thời gian', 'ten_hang': 'Tên hàng',
            'ngay_du_kien': 'Ngày dự kiến giao',
            'quy_cach': 'Quy cách', 'so_luong': 'Số lượng',
            'mau_sac': 'Màu sắc', 'mau_keo': 'Màu keo',
            'don_gia_goc': 'Đơn giá gốc', 'thanh_tien': 'Thành tiền',
            'don_gia_ban': 'Đơn giá bán', 'thanh_tien_ban': 'Thành tiền bán',
            'cong_no_khach': 'Công nợ khách', 'ctv': 'CTV',
            'hoa_hong': 'Hoa hồng', 'tien_hoa_hong': 'Tiền hoa hồng',
            'loi_nhuan': 'Lợi nhuận',
            'da_giao': 'Đã giao',
            'da_tat_toan': 'Đã tất toán'
        }
        
        for col in columns:
            self.truc_in_tree.heading(col, text=headings[col])
            # Hide ID column
            if col == 'id':
                self.truc_in_tree.column(col, width=50, stretch=False)
            else:
                # Adjust column widths based on content
                width = max(len(headings[col])*10, 100)
                self.truc_in_tree.column(col, width=width, minwidth=50)
        
        # Add scrollbars
        y_scrollbar = ttk.Scrollbar(container, orient=tk.VERTICAL, command=self.truc_in_tree.yview)
        x_scrollbar = ttk.Scrollbar(container, orient=tk.HORIZONTAL, command=self.truc_in_tree.xview)
        self.truc_in_tree.configure(yscrollcommand=y_scrollbar.set, xscrollcommand=x_scrollbar.set)
        
        # Grid layout for better performance
        self.truc_in_tree.grid(row=0, column=0, sticky='nsew')
        y_scrollbar.grid(row=0, column=1, sticky='ns')
        x_scrollbar.grid(row=1, column=0, sticky='ew')
        
        # Bind double-click event with delay
        self.truc_in_tree.bind('<Double-1>', lambda e: self.root.after(100, lambda: self.show_edit_form('truc_in')))
        
        # Bind Ctrl+A for select all
        self.truc_in_tree.bind('<Control-a>', lambda e: self.select_all(self.truc_in_tree))
        
    def select_all(self, tree):
        """Select all items in the treeview"""
        tree.selection_set(tree.get_children())
        return 'break'  # Prevent default Ctrl+A behavior
        
    def on_tab_changed(self, event):
        # Hide edit frame when switching tabs
        self.edit_frame.pack_forget()
        
    def show_edit_form(self, order_type):
        # Get selected item
        if order_type == 'bang_keo_in':
            tree = self.bang_keo_in_tree
            # Define read-only fields for bang_keo_in
            readonly_fields = [
                'id',  # Add ID to readonly fields
                'don_gia_goc', 'thanh_tien_goc', 'thanh_tien_ban',
                'cong_no_khach', 'tien_hoa_hong', 'loi_nhuan'
            ]
            # Fields that trigger recalculation
            calculation_trigger_fields = [
                'so_luong', 'phi_sl', 'phi_keo', 'phi_size', 'phi_cat', 'don_gia_von', 'don_gia_ban', 'tien_coc',
                'hoa_hong', 'phi_mau', 'quy_cach_mm', 'quy_cach_m', 'quy_cach_mic', 'cuon_cay'
            ]
            window_title = "Chỉnh sửa đơn hàng băng keo"
        else:
            tree = self.truc_in_tree
            # Define read-only fields for truc_in
            readonly_fields = [
                'id',  # Add ID to readonly fields
                'thanh_tien', 'thanh_tien_ban', 'cong_no_khach',
                'tien_hoa_hong', 'loi_nhuan'
            ]
            # Fields that trigger recalculation
            calculation_trigger_fields = [
                'so_luong', 'don_gia_ban', 'don_gia_goc', 'hoa_hong'
            ]
            window_title = "Chỉnh sửa đơn hàng trục in"
            
        selected_item = tree.selection()
        if not selected_item:
            return
            
        # Get values
        values = tree.item(selected_item)['values']
        self.current_edit_item = selected_item[0]
        self.current_edit_type = order_type
        
        # Create a new top-level window
        edit_window = tk.Toplevel(self.root)
        edit_window.title(window_title)
        edit_window.geometry("800x600")
        edit_window.minsize(800, 600)
        
        # Make the window modal
        edit_window.transient(self.root)
        edit_window.grab_set()
        
        # Center the window on screen
        window_width = 800
        window_height = 600
        screen_width = edit_window.winfo_screenwidth()
        screen_height = edit_window.winfo_screenheight()
        center_x = int((screen_width - window_width) / 2)
        center_y = int((screen_height - window_height) / 2)
        edit_window.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
        
        # Create main frame with padding
        main_frame = ttk.Frame(edit_window, padding="10 10 10 10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        self.edit_entries = {}
        
        if order_type == 'bang_keo_in':
            # Create frames for different sections
            info_frame = ttk.LabelFrame(main_frame, text="Thông tin đơn hàng", padding="5 5 5 5")
            info_frame.pack(fill=tk.X, padx=5, pady=5)
            
            specs_frame = ttk.LabelFrame(main_frame, text="Quy cách", padding="5 5 5 5")
            specs_frame.pack(fill=tk.X, padx=5, pady=5)
            
            price_frame = ttk.LabelFrame(main_frame, text="Giá và phí", padding="5 5 5 5")
            price_frame.pack(fill=tk.X, padx=5, pady=5)
            
            # Info section
            self._create_field(info_frame, 'thoi_gian', 'Thời gian:', values, 0, 0, readonly_fields)
            self._create_field(info_frame, 'ten_hang', 'Tên hàng:', values, 0, 2, readonly_fields)
            self._create_field(info_frame, 'ngay_du_kien', 'Ngày dự kiến:', values, 1, 0, readonly_fields)
            
            # Specs section
            self._create_field(specs_frame, 'quy_cach_mm', 'Quy cách (mm):', values, 0, 0, readonly_fields)
            self._create_field(specs_frame, 'quy_cach_m', 'Quy cách (m):', values, 0, 2, readonly_fields)
            self._create_field(specs_frame, 'quy_cach_mic', 'Quy cách (mic):', values, 1, 0, readonly_fields)
            self._create_field(specs_frame, 'cuon_cay', 'Cuộn/cây:', values, 1, 2, readonly_fields)
            self._create_field(specs_frame, 'so_luong', 'Số lượng:', values, 2, 0, readonly_fields)
            
            # Price section
            self._create_field(price_frame, 'phi_sl', 'Phí SL:', values, 0, 0, readonly_fields)
            self._create_field(price_frame, 'mau_keo', 'Màu keo:', values, 0, 2, readonly_fields)
            self._create_field(price_frame, 'phi_keo', 'Phí keo:', values, 0, 4, readonly_fields)
            
            self._create_field(price_frame, 'mau_sac', 'Màu sắc:', values, 1, 0, readonly_fields)
            self._create_field(price_frame, 'phi_mau', 'Phí màu:', values, 1, 2, readonly_fields)
            self._create_field(price_frame, 'phi_size', 'Phí size:', values, 1, 4, readonly_fields)
            
            self._create_field(price_frame, 'phi_cat', 'Phí cắt:', values, 2, 0, readonly_fields)
            self._create_field(price_frame, 'don_gia_von', 'Đơn giá vốn:', values, 2, 2, readonly_fields)
            self._create_field(price_frame, 'don_gia_goc', 'Đơn giá gốc:', values, 2, 4, readonly_fields)
            
            self._create_field(price_frame, 'thanh_tien_goc', 'Thành tiền gốc:', values, 3, 0, readonly_fields)
            self._create_field(price_frame, 'don_gia_ban', 'Đơn giá bán:', values, 3, 2, readonly_fields)
            self._create_field(price_frame, 'thanh_tien_ban', 'Thành tiền bán:', values, 3, 4, readonly_fields)
            
            self._create_field(price_frame, 'tien_coc', 'Tiền cọc:', values, 4, 0, readonly_fields)
            self._create_field(price_frame, 'cong_no_khach', 'Công nợ khách:', values, 4, 2, readonly_fields)
            
            self._create_field(price_frame, 'ctv', 'CTV:', values, 5, 0, readonly_fields)
            self._create_field(price_frame, 'hoa_hong', 'Hoa hồng (%):', values, 5, 2, readonly_fields)
            self._create_field(price_frame, 'tien_hoa_hong', 'Tiền hoa hồng:', values, 5, 4, readonly_fields)
            
            self._create_field(price_frame, 'loi_giay', 'Lõi giấy:', values, 6, 0, readonly_fields)
            self._create_field(price_frame, 'thung_bao', 'Thùng/Bao:', values, 6, 2, readonly_fields)
            self._create_field(price_frame, 'loi_nhuan', 'Lợi nhuận:', values, 6, 4, readonly_fields)

        else:  # truc_in
            # Create frames for different sections
            info_frame = ttk.LabelFrame(main_frame, text="Thông tin đơn hàng", padding="5 5 5 5")
            info_frame.pack(fill=tk.X, padx=5, pady=5)
            
            specs_frame = ttk.LabelFrame(main_frame, text="Quy cách", padding="5 5 5 5")
            specs_frame.pack(fill=tk.X, padx=5, pady=5)
            
            price_frame = ttk.LabelFrame(main_frame, text="Giá và phí", padding="5 5 5 5")
            price_frame.pack(fill=tk.X, padx=5, pady=5)
            
            # Info section
            self._create_field(info_frame, 'thoi_gian', 'Thời gian:', values, 0, 0, readonly_fields)
            self._create_field(info_frame, 'ten_hang', 'Tên hàng:', values, 0, 2, readonly_fields)
            self._create_field(info_frame, 'ngay_du_kien', 'Ngày dự kiến:', values, 1, 0, readonly_fields)
            
            # Specs section
            self._create_field(specs_frame, 'quy_cach', 'Quy cách:', values, 0, 0, readonly_fields)
            self._create_field(specs_frame, 'so_luong', 'Số lượng:', values, 0, 2, readonly_fields)
            self._create_field(specs_frame, 'mau_sac', 'Màu sắc:', values, 1, 0, readonly_fields)
            self._create_field(specs_frame, 'mau_keo', 'Màu keo:', values, 1, 2, readonly_fields)
            
            # Price section
            self._create_field(price_frame, 'don_gia_goc', 'Đơn giá gốc:', values, 0, 0, readonly_fields)
            self._create_field(price_frame, 'thanh_tien', 'Thành tiền:', values, 0, 2, readonly_fields)
            
            self._create_field(price_frame, 'don_gia_ban', 'Đơn giá bán:', values, 1, 0, readonly_fields)
            self._create_field(price_frame, 'thanh_tien_ban', 'Thành tiền bán:', values, 1, 2, readonly_fields)
            
            self._create_field(price_frame, 'cong_no_khach', 'Công nợ khách:', values, 2, 0, readonly_fields)
            
            self._create_field(price_frame, 'ctv', 'CTV:', values, 3, 0, readonly_fields)
            self._create_field(price_frame, 'hoa_hong', 'Hoa hồng (%):', values, 3, 2, readonly_fields)
            
            self._create_field(price_frame, 'tien_hoa_hong', 'Tiền hoa hồng:', values, 4, 0, readonly_fields)
            self._create_field(price_frame, 'loi_nhuan', 'Lợi nhuận:', values, 4, 2, readonly_fields)
        
        # Add button frame at the bottom
        button_frame = ttk.Frame(edit_window)
        button_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=10)
        
        # Add save and cancel buttons
        ttk.Button(button_frame, text="Lưu", command=lambda: self.save_edit_and_close(edit_window)).pack(side=tk.RIGHT, padx=5)
        ttk.Button(button_frame, text="Hủy", command=edit_window.destroy).pack(side=tk.RIGHT, padx=5)
        
        # Bind keyboard shortcuts
        edit_window.bind('<Escape>', lambda e: edit_window.destroy())
        edit_window.bind('<Control-s>', lambda e: self.save_edit_and_close(edit_window))
        
        # Focus the first non-readonly entry
        for entry in self.edit_entries.values():
            if entry.cget('state') != 'readonly':
                entry.focus_set()
                break

    def _create_field(self, parent, field_name, label_text, values, row, col, readonly_fields):
        """Helper method to create a labeled field"""
        ttk.Label(parent, text=label_text).grid(row=row, column=col, sticky=tk.W, padx=5, pady=5)
        
        # Get the correct index for the value based on the field name
        field_index = list(self.bang_keo_in_tree['columns'] if self.current_edit_type == 'bang_keo_in' else self.truc_in_tree['columns']).index(field_name)
        
        # Special handling for date fields
        if field_name in ['thoi_gian', 'ngay_du_kien']:
            date_value = values[field_index] if values[field_index] else datetime.now().strftime(self.DATE_FORMAT)
            entry = DateEntry(parent, width=15, background='darkblue',
                             foreground='white', borderwidth=2,
                             date_pattern='dd/mm/yyyy',
                             locale='vi_VN')
            try:
                entry.set_date(datetime.strptime(date_value, self.DATE_FORMAT).date())
            except ValueError:
                try:
                    # Try parsing with datetime format if date format fails
                    entry.set_date(datetime.strptime(date_value, self.DATETIME_FORMAT).date())
                except ValueError:
                    entry.set_date(datetime.now().date())
        else:
            entry = ttk.Entry(parent, width=15)
            entry.insert(0, values[field_index] if values[field_index] is not None else '')
            
            # Make field read-only if it's in readonly_fields
            if field_name in readonly_fields:
                entry.configure(state='readonly')
            
            # Bind calculation event if field triggers calculation
            calculation_trigger_fields = {
                'bang_keo_in': [
                    'so_luong', 'phi_sl', 'phi_keo', 'phi_size', 'phi_cat', 'don_gia_von',
                    'don_gia_ban', 'tien_coc', 'hoa_hong', 'phi_mau', 'quy_cach_mm',
                    'quy_cach_m', 'quy_cach_mic', 'cuon_cay'
                ],
                'truc_in': [
                    'so_luong', 'don_gia_ban', 'don_gia_goc', 'hoa_hong'
                ]
            }[self.current_edit_type]
            
            if field_name in calculation_trigger_fields:
                entry.bind('<KeyRelease>', lambda e, ot=self.current_edit_type: self.recalculate_edit_form(ot))
                entry.bind('<FocusOut>', self.format_currency_input)
        
        entry.grid(row=row, column=col+1, sticky=tk.W, padx=5, pady=5)
        self.edit_entries[field_name] = entry

    def _bind_mousewheel(self, canvas):
        """Bind mousewheel to scrolling"""
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        
    def _unbind_mousewheel(self, canvas):
        """Unbind mousewheel from scrolling"""
        canvas.unbind_all("<MouseWheel>")

    def save_edit_and_close(self, edit_window):
        """Save the edit and close the window"""
        self.save_edit()
        edit_window.destroy()

    def recalculate_edit_form(self, order_type):
        """Recalculate values in the edit form based on the order type"""
        try:
            if order_type == 'bang_keo_in':
                # Get values from edit entries
                don_gia_von = self.validate_float_input(self.edit_entries['don_gia_von'].get())
                phi_sl = self.validate_float_input(self.edit_entries['phi_sl'].get())
                phi_mau = self.validate_float_input(self.edit_entries['phi_mau'].get())
                phi_keo = self.validate_float_input(self.edit_entries['phi_keo'].get())
                phi_size = self.validate_float_input(self.edit_entries['phi_size'].get())
                phi_cat = self.validate_float_input(self.edit_entries['phi_cat'].get())
                so_luong = self.validate_float_input(self.edit_entries['so_luong'].get())
                don_gia_ban = self.validate_float_input(self.edit_entries['don_gia_ban'].get())
                tien_coc = self.validate_float_input(self.edit_entries['tien_coc'].get())
                hoa_hong = self.validate_float_input(self.edit_entries['hoa_hong'].get()) / 100
                cuon_cay = self.validate_float_input(self.edit_entries['cuon_cay'].get())
                quy_cach_m = self.validate_float_input(self.edit_entries['quy_cach_m'].get())

                # Calculate don_gia_goc
                if cuon_cay == 0 or quy_cach_m == 0:
                    don_gia_goc = 0
                else:
                    don_gia_goc = (don_gia_von + phi_sl + phi_mau + phi_keo + phi_size + phi_cat) / 90 * quy_cach_m / cuon_cay

                # Calculate other values
                thanh_tien_goc = don_gia_goc * so_luong
                thanh_tien_ban = don_gia_ban * so_luong
                cong_no_khach = thanh_tien_ban - tien_coc
                loi_nhuan = thanh_tien_ban - thanh_tien_goc
                tien_hoa_hong = loi_nhuan * hoa_hong

                # Update readonly fields
                self.update_readonly_field(self.edit_entries['don_gia_goc'], don_gia_goc)
                self.update_readonly_field(self.edit_entries['thanh_tien_goc'], thanh_tien_goc)
                self.update_readonly_field(self.edit_entries['thanh_tien_ban'], thanh_tien_ban)
                self.update_readonly_field(self.edit_entries['cong_no_khach'], cong_no_khach)
                self.update_readonly_field(self.edit_entries['tien_hoa_hong'], tien_hoa_hong)
                self.update_readonly_field(self.edit_entries['loi_nhuan'], loi_nhuan)

            else:  # truc_in
                # Get values
                so_luong = self.validate_float_input(self.edit_entries['so_luong'].get())
                don_gia_ban = self.validate_float_input(self.edit_entries['don_gia_ban'].get())
                don_gia_goc = self.validate_float_input(self.edit_entries['don_gia_goc'].get())
                hoa_hong = self.validate_float_input(self.edit_entries['hoa_hong'].get()) / 100

                # Calculate values
                thanh_tien = don_gia_goc * so_luong
                thanh_tien_ban = don_gia_ban * so_luong
                loi_nhuan = thanh_tien_ban - thanh_tien
                tien_hoa_hong = loi_nhuan * hoa_hong
                cong_no_khach = thanh_tien_ban

                # Update readonly fields
                self.update_readonly_field(self.edit_entries['thanh_tien'], thanh_tien)
                self.update_readonly_field(self.edit_entries['thanh_tien_ban'], thanh_tien_ban)
                self.update_readonly_field(self.edit_entries['cong_no_khach'], cong_no_khach)
                self.update_readonly_field(self.edit_entries['tien_hoa_hong'], tien_hoa_hong)
                self.update_readonly_field(self.edit_entries['loi_nhuan'], loi_nhuan)

        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi tính toán: {str(e)}")
        
    def save_edit(self):
        if not self.current_edit_item or not self.current_edit_type:
            return
            
        try:
            # Get values from entries
            values = {}
            tree = self.bang_keo_in_tree if self.current_edit_type == 'bang_keo_in' else self.truc_in_tree
            
            # Get the record ID from the hidden first column
            item_values = tree.item(self.current_edit_item)['values']
            record_id = item_values[0]  # Get ID from the first column
            
            # Get all values from edit entries
            for column in tree['columns']:
                if column in ['id', 'da_giao', 'da_tat_toan']:  # Skip these fields
                    continue
                if column in ['thoi_gian', 'ngay_du_kien']:
                    # Get date from DateEntry widget and format it
                    date_value = self.edit_entries[column].get_date()
                    values[column] = date_value.strftime(self.DATE_FORMAT)
                else:
                    if column in self.edit_entries:  # Only get values for fields that exist in edit form
                        values[column] = self.edit_entries[column].get()
            
            # Update database
            try:
                if self.current_edit_type == 'bang_keo_in':
                    order = self.db_session.query(BangKeoInOrder).filter_by(id=record_id).first()
                    
                    if order:
                        for field, value in values.items():
                            if field not in ['id', 'da_giao', 'da_tat_toan']:
                                if field in ['thoi_gian', 'ngay_du_kien']:
                                    # Convert string date to datetime for database
                                    setattr(order, field, datetime.strptime(value, self.DATE_FORMAT))
                                elif field in ['quy_cach_mm', 'quy_cach_m', 'quy_cach_mic', 'cuon_cay', 'so_luong', 
                                          'phi_sl', 'phi_keo', 'phi_mau', 'phi_size', 'phi_cat', 'don_gia_von', 
                                          'don_gia_goc', 'thanh_tien_goc', 'don_gia_ban', 'thanh_tien_ban', 
                                          'tien_coc', 'cong_no_khach', 'hoa_hong', 'tien_hoa_hong', 'loi_nhuan']:
                                    setattr(order, field, self.validate_float_input(value))
                                else:
                                    setattr(order, field, value)
                else:
                    order = self.db_session.query(TrucInOrder).filter_by(id=record_id).first()
                    
                    if order:
                        for field, value in values.items():
                            if field not in ['id', 'da_giao', 'da_tat_toan']:
                                if field in ['thoi_gian', 'ngay_du_kien']:
                                    setattr(order, field, datetime.strptime(value, self.DATE_FORMAT))
                                elif field in ['so_luong', 'don_gia_goc', 'thanh_tien', 'don_gia_ban', 
                                          'thanh_tien_ban', 'cong_no_khach', 'hoa_hong', 'tien_hoa_hong', 'loi_nhuan']:
                                    setattr(order, field, self.validate_float_input(value))
                                else:
                                    setattr(order, field, value)
                
                if not order:
                    raise Exception("Không tìm thấy đơn hàng trong database")
                
                # Commit changes to database
                self.db_session.commit()
                
                # Update tree with new values
                tree_values = [record_id]  # Start with the ID
                for column in tree['columns'][1:]:  # Skip the ID column
                    if column in ['thoi_gian', 'ngay_du_kien']:
                        tree_values.append(values[column])
                    elif column in ['da_giao', 'da_tat_toan']:
                        tree_values.append(item_values[tree['columns'].index(column)])
                    else:
                        tree_values.append(values.get(column, ''))
                
                # Update tree
                tree.item(self.current_edit_item, values=tree_values)
                
                # Update stored items list
                if self.current_edit_type == 'bang_keo_in':
                    for i, (item_id, _) in enumerate(self.all_bang_keo_in_items):
                        if item_id == self.current_edit_item:
                            self.all_bang_keo_in_items[i] = (item_id, tree_values)
                            break
                else:
                    for i, (item_id, _) in enumerate(self.all_truc_in_items):
                        if item_id == self.current_edit_item:
                            self.all_truc_in_items[i] = (item_id, tree_values)
                            break
                
                messagebox.showinfo("Thành công", "Đã cập nhật thông tin")
                
            except Exception as e:
                self.db_session.rollback()
                raise Exception(f"Lỗi khi cập nhật database: {str(e)}")
                
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi cập nhật: {str(e)}")

    def add_order(self, order_type, order_data):
        """Add a new order to the history"""
        current_time = datetime.now().strftime(self.DATE_FORMAT)
        
        # Helper function to format currency fields
        def format_value(key, value):
            currency_fields = [
                'phi_sl', 'phi_keo', 'phi_mau', 'phi_size', 'phi_cat',
                'don_gia_von', 'don_gia_goc', 'thanh_tien_goc', 'don_gia_ban',
                'thanh_tien_ban', 'tien_coc', 'cong_no_khach', 'tien_hoa_hong',
                'loi_nhuan', 'thanh_tien'
            ]
            return self.format_currency(value) if key in currency_fields else value
        
        if order_type == 'Băng keo in':
            values = [
                order_data.get('id', ''),
                order_data.get('thoi_gian', current_time),
                order_data.get('ten_hang', ''),
                order_data.get('ngay_du_kien', current_time),
                order_data.get('quy_cach_mm', ''),
                order_data.get('quy_cach_m', ''),
                order_data.get('quy_cach_mic', ''),
                order_data.get('cuon_cay', ''),
                order_data.get('so_luong', ''),
                format_value('phi_sl', order_data.get('phi_sl', '')),
                order_data.get('mau_keo', ''),
                format_value('phi_keo', order_data.get('phi_keo', '')),
                order_data.get('mau_sac', ''),
                format_value('phi_mau', order_data.get('phi_mau', '')),
                format_value('phi_size', order_data.get('phi_size', '')),
                format_value('phi_cat', order_data.get('phi_cat', '')),
                format_value('don_gia_von', order_data.get('don_gia_von', '')),
                format_value('don_gia_goc', order_data.get('don_gia_goc', '')),
                format_value('thanh_tien_goc', order_data.get('thanh_tien_goc', '')),
                format_value('don_gia_ban', order_data.get('don_gia_ban', '')),
                format_value('thanh_tien_ban', order_data.get('thanh_tien_ban', '')),
                format_value('tien_coc', order_data.get('tien_coc', '')),
                format_value('cong_no_khach', order_data.get('cong_no_khach', '')),
                order_data.get('ctv', ''),
                order_data.get('hoa_hong', ''),
                format_value('tien_hoa_hong', order_data.get('tien_hoa_hong', '')),
                order_data.get('loi_giay', ''),
                order_data.get('thung_bao', ''),
                format_value('loi_nhuan', order_data.get('loi_nhuan', '')),
                order_data.get('da_giao', ''),
                order_data.get('da_tat_toan', '')
            ]
            item = self.bang_keo_in_tree.insert('', 0, values=values)
            self.bang_keo_in_data.append(item)
            self.all_bang_keo_in_items.append((item, values))
        else:  # Trục in
            values = [
                order_data.get('id', ''),
                order_data.get('thoi_gian', current_time),
                order_data.get('ten_hang', ''),
                order_data.get('ngay_du_kien', current_time),
                order_data.get('quy_cach', ''),
                order_data.get('so_luong', ''),
                order_data.get('mau_sac', ''),
                order_data.get('mau_keo', ''),
                format_value('don_gia_goc', order_data.get('don_gia_goc', '')),
                format_value('thanh_tien', order_data.get('thanh_tien', '')),
                format_value('don_gia_ban', order_data.get('don_gia_ban', '')),
                format_value('thanh_tien_ban', order_data.get('thanh_tien_ban', '')),
                format_value('cong_no_khach', order_data.get('cong_no_khach', '')),
                order_data.get('ctv', ''),
                order_data.get('hoa_hong', ''),
                format_value('tien_hoa_hong', order_data.get('tien_hoa_hong', '')),
                format_value('loi_nhuan', order_data.get('loi_nhuan', '')),
                order_data.get('da_giao', ''),
                order_data.get('da_tat_toan', '')
            ]
            item = self.truc_in_tree.insert('', 0, values=values)
            self.truc_in_data.append(item)
            self.all_truc_in_items.append((item, values))
        
    def export_selected_to_excel(self):
        try:
            # Get current tab and tree
            current_tab = self.category_notebook.select()
            if current_tab == str(self.bang_keo_in_frame):
                tree = self.bang_keo_in_tree
                sheet_name = "Bang keo in"
            else:
                tree = self.truc_in_tree
                sheet_name = "Truc in"
            
            # Get selected items
            selected_items = tree.selection()
            if not selected_items:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn ít nhất một dòng để xuất Excel")
                return
            
            # Get column headers
            headers = [tree.heading(col)['text'] for col in tree['columns']]
            
            # Get data from selected items and convert dates
            data = []
            for item in selected_items:
                values = list(tree.item(item)['values'])
                # Convert date formats for relevant columns
                if current_tab == str(self.bang_keo_in_frame):
                    # For bang_keo_in_tree: thoi_gian is index 1, ngay_du_kien is index 3
                    if values[1]:  # thoi_gian
                        try:
                            # Parse the original date string and convert to MM/DD/YYYY
                            date_obj = datetime.strptime(values[1], self.DATE_FORMAT)
                            values[1] = date_obj.strftime('%m/%d/%Y')
                        except ValueError:
                            pass
                    if values[3]:  # ngay_du_kien
                        try:
                            date_obj = datetime.strptime(values[3], self.DATE_FORMAT)
                            values[3] = date_obj.strftime('%m/%d/%Y')
                        except ValueError:
                            pass
                else:
                    # For truc_in_tree: thoi_gian is index 1, ngay_du_kien is index 3
                    if values[1]:  # thoi_gian
                        try:
                            date_obj = datetime.strptime(values[1], self.DATE_FORMAT)
                            values[1] = date_obj.strftime('%m/%d/%Y')
                        except ValueError:
                            pass
                    if values[3]:  # ngay_du_kien
                        try:
                            date_obj = datetime.strptime(values[3], self.DATE_FORMAT)
                            values[3] = date_obj.strftime('%m/%d/%Y')
                        except ValueError:
                            pass
                data.append(values)
            
            # Get save file location
            current_time = datetime.now().strftime('%d-%m-%y')
            default_filename = f"DonHang_{sheet_name}_{current_time}.xlsx"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=default_filename
            )
            
            if not file_path:
                return

            # Create or load workbook
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    next_row = ws.max_row + 1
                else:
                    ws = wb.create_sheet(sheet_name)
                    # Write headers for new sheet
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)
                    next_row = 2
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name
                # Write headers
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                next_row = 2

            # Write data to rows
            for row_data in data:
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=next_row, column=col, value=value)
                next_row += 1

            # Save file
            wb.save(file_path)
            
            messagebox.showinfo("Thành công", f"Đã xuất dữ liệu ra file Excel:\n{file_path}")
            self.update_status("Đã xuất Excel thành công")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi xuất Excel: {str(e)}")
            self.update_status("Lỗi khi xuất Excel")

    def load_bang_keo_in_data(self):
        """Load data for bang keo treeview"""
        try:
            # Store current items before clearing
            current_items = []
            for item in self.bang_keo_in_tree.get_children():
                values = self.bang_keo_in_tree.item(item)['values']
                current_items.append(values)
                
            # Clear existing items
            for item in self.bang_keo_in_tree.get_children():
                self.bang_keo_in_tree.delete(item)
                
            # Restore the items
            for values in current_items:
                self.bang_keo_in_tree.insert('', 'end', values=values)
                
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi làm mới dữ liệu: {str(e)}")

    def load_truc_in_data(self):
        """Load data for truc in treeview"""
        try:
            # Store current items before clearing
            current_items = []
            for item in self.truc_in_tree.get_children():
                values = self.truc_in_tree.item(item)['values']
                current_items.append(values)
                
            # Clear existing items
            for item in self.truc_in_tree.get_children():
                self.truc_in_tree.delete(item)
                
            # Restore the items
            for values in current_items:
                self.truc_in_tree.insert('', 'end', values=values)
                
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi làm mới dữ liệu: {str(e)}")

    def apply_filters(self):
        """Apply date and search filters"""
        try:
            from_date = self.from_date.get_date()
            to_date = self.to_date.get_date()
            search_text = self.search_var.get().lower()
            
            # Get current tab and tree
            current_tab = self.category_notebook.select()
            if current_tab == str(self.bang_keo_in_frame):
                tree = self.bang_keo_in_tree
                all_items = self.all_bang_keo_in_items
            else:
                tree = self.truc_in_tree
                all_items = self.all_truc_in_items
            
            # Clear current view
            for item in tree.get_children():
                tree.delete(item)
            
            # Reinsert items that match the filter
            for item_id, values in all_items:
                try:
                    # Check search text if present
                    text_match = True
                    if search_text:
                        text_match = False
                        for value in values:
                            if str(value).lower().find(search_text) != -1:
                                text_match = True
                                break
                    
                    # Only check dates if text matches and dates are selected
                    if text_match and from_date and to_date:
                        # Parse the order date (thoi_gian - index 1)
                        order_date_str = values[1]
                        order_date = datetime.strptime(order_date_str, self.DATE_FORMAT).date()
                        
                        # Check if order date is within range
                        date_match = from_date <= order_date <= to_date
                        text_match = text_match and date_match
                    
                    # Insert if conditions are met
                    if text_match:
                        tree.insert('', 'end', values=values)
                        
                        # Apply alternating row colors
                        items = tree.get_children()
                        for i, item in enumerate(items):
                            if i % 2 == 0:
                                tree.tag_configure('evenrow', background='#FFFFFF')
                                tree.item(item, tags=('evenrow',))
                            else:
                                tree.tag_configure('oddrow', background='#F0F0F0')
                                tree.item(item, tags=('oddrow',))
                
                except (ValueError, IndexError) as e:
                    logging.error(f"Error processing item in apply_filters: {str(e)}")
                    continue
            
            self.update_status("Đã áp dụng bộ lọc")
            
        except Exception as e:
            logging.error(f"Error in apply_filters: {str(e)}")
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi lọc dữ liệu: {str(e)}")

    def reset_filters(self):
        """Reset all filters"""
        try:
            # Reset date pickers to current date
            current_date = datetime.now().date()
            self.from_date.set_date(current_date)
            self.to_date.set_date(current_date)
            
            # Clear search
            self.search_var.set('')
            
            # Get current tab and tree
            current_tab = self.category_notebook.select()
            if current_tab == str(self.bang_keo_in_frame):
                tree = self.bang_keo_in_tree
                all_items = self.all_bang_keo_in_items
            else:
                tree = self.truc_in_tree
                all_items = self.all_truc_in_items
            
            # Clear current view
            for item in tree.get_children():
                tree.delete(item)
            
            # Restore all items
            for _, values in all_items:  # Changed from item_id, values to _, values
                tree.insert('', 'end', values=values)  # Removed item_id
            
            messagebox.showinfo("Thông báo", "Đã đặt lại bộ lọc")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi đặt lại bộ lọc: {str(e)}")

    def edit_item(self, event=None):
        """Handle double-click to edit item"""
        current_tab = self.category_notebook.select()
        if current_tab == str(self.bang_keo_in_frame):
            tree = self.bang_keo_in_tree
            item_type = 'bang_keo_in'
        else:
            tree = self.truc_in_tree
            item_type = 'truc_in'

        selected_items = tree.selection()
        if not selected_items:
            return

        item = selected_items[0]  # Get the first selected item
        values = tree.item(item)['values']
        
        # Show edit dialog
        self.show_edit_dialog(item, values, item_type)

    def show_edit_dialog(self, item, values, item_type):
        """Show dialog to edit item"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Chỉnh sửa thông tin")
        dialog.transient(self.root)
        dialog.grab_set()

        # Center the dialog
        dialog.geometry("400x300")
        dialog.resizable(False, False)

        # Create and pack widgets
        main_frame = ttk.Frame(dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Create entry fields based on item type
        entries = {}
        row = 0
        if item_type == 'bang_keo_in':
            fields = ['ten_hang', 'so_luong', 'don_gia_ban', 'tien_coc']
            labels = ['Tên hàng:', 'Số lượng:', 'Đơn giá bán:', 'Tiền cọc:']
        else:  # truc_in
            fields = ['ten_hang', 'so_luong', 'don_gia_ban']
            labels = ['Tên hàng:', 'Số lượng:', 'Đơn giá bán:']

        for field, label in zip(fields, labels):
            ttk.Label(main_frame, text=label).grid(row=row, column=0, sticky='e', padx=5, pady=5)
            entry = ttk.Entry(main_frame, width=30)
            entry.grid(row=row, column=1, sticky='w', padx=5, pady=5)
            entry.insert(0, values[fields.index(field) + 1])  # +1 because first column is timestamp
            entries[field] = entry
            row += 1

        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=row, column=0, columnspan=2, pady=20)

        def save_changes():
            # Get values from entries
            new_values = {}
            for field, entry in entries.items():
                new_values[field] = entry.get()

            # Update database
            try:
                if item_type == 'bang_keo_in':
                    order = self.db_session.query(BangKeoInOrder).filter_by(
                        thoi_gian=datetime.strptime(values['thoi_gian'], '%d/%m/%Y %H:%M:%S')
                    ).first()
                else:
                    order = self.db_session.query(TrucInOrder).filter_by(
                        thoi_gian=datetime.strptime(values['thoi_gian'], '%d/%m/%Y %H:%M:%S')
                    ).first()

                if order:
                    for field, value in new_values.items():
                        setattr(order, field, value)
                    self.db_session.commit()

                    # Update treeview
                    tree = self.bang_keo_in_tree if item_type == 'bang_keo_in' else self.truc_in_tree
                    new_values_list = list(values)
                    for i, field in enumerate(fields):
                        new_values_list[i + 1] = new_values[field]
                    tree.item(item, values=new_values_list)

                    messagebox.showinfo("Thành công", "Đã cập nhật thông tin")
                    dialog.destroy()
                    self.update_status("Đã cập nhật thông tin")

            except Exception as e:
                self.db_session.rollback()
                messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi cập nhật: {str(e)}")
                self.update_status("Lỗi khi cập nhật thông tin")

        ttk.Button(button_frame, text="Lưu", command=save_changes).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Hủy", command=dialog.destroy).pack(side=tk.LEFT, padx=5)

        # Bind double-click event
        self.bang_keo_in_tree.bind('<Double-1>', self.edit_item)
        self.truc_in_tree.bind('<Double-1>', self.edit_item)

    def create_import_export_buttons(self):
        # Create a frame for buttons
        button_frame = ttk.Frame(self.tab)
        button_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)
        
        # Configure grid to make buttons stick during resizing
        button_frame.columnconfigure((0, 1, 2, 3), weight=1)
    
        # Export Template Buttons
        export_bang_keo_in_button = ttk.Button(button_frame, text="Export Băng Keo Template", command=lambda: self.export_template('bang_keo_in'))
        export_bang_keo_in_button.grid(row=0, column=0, padx=5, sticky='ew')
    
        export_truc_in_button = ttk.Button(button_frame, text="Export Trục In Template", command=lambda: self.export_template('truc_in'))
        export_truc_in_button.grid(row=0, column=1, padx=5, sticky='ew')
    
        # Import Data Buttons
        import_bang_keo_in_button = ttk.Button(button_frame, text="Import Băng Keo Data", command=lambda: self.import_data('bang_keo_in'))
        import_bang_keo_in_button.grid(row=0, column=2, padx=5, sticky='ew')
    
        import_truc_in_button = ttk.Button(button_frame, text="Import Trục In Data", command=lambda: self.import_data('truc_in'))
        import_truc_in_button.grid(row=0, column=3, padx=5, sticky='ew')

    def export_template(self, order_type):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            export_template(file_path, order_type)

    def import_data(self, order_type):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            import_data(file_path, order_type, self.db_session)
            self.refresh_data()  # Refresh data to show imported entries
            
    def export_selected_to_email(self):
        """Export selected order details to a text file and open it."""
        try:
            current_tab = self.category_notebook.select()
            if current_tab == str(self.bang_keo_in_frame):
                tree = self.bang_keo_in_tree
            else:
                tree = self.truc_in_tree

            selected_items = tree.selection()
            if not selected_items:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn ít nhất một dòng để xuất email")
                return

            # Get the first selected item's values
            values = tree.item(selected_items[0])['values']
            
            # Create email content based on the order type
            if current_tab == str(self.bang_keo_in_frame):
                # Format for bang keo in
                email_content = (
                    f"Chào bác,\n\n"
                    f"Bác làm giúp con đơn hàng in logo \"{values[2]}\" này nhé\n"  # ten_hang
                    f"Màu sắc: {values[12]} / Màu keo: {values[10]}\n"  # mau_sac, mau_keo
                    f"Số lượng: {values[8]} cuộn\n"  # so_luong
                    f"Quy cách: {values[4]}mm * {values[5]}m * {values[6]}mic\n"  # quy_cach
                    f"Lõi giấy: {values[26]} - Thùng bao: {values[27]}\n\n"  # loi_giay, thung_bao
                    f"Cám ơn bác\n"
                    f"Quế"
                )
            else:
                # Format for truc in
                email_content = (
                    f"Chào bác,\n\n"
                    f"Bác làm giúp con đơn hàng trục in \"{values[2]}\" này nhé\n"  # ten_hang
                    f"Màu sắc: {values[6]} / Màu keo: {values[7]}\n"  # mau_sac, mau_keo
                    f"Số lượng: {values[5]} cái\n"  # so_luong
                    f"Quy cách: {values[4]}mm\n\n"  # quy_cach
                    f"Cám ơn bác\n"
                    f"Quế"
                )

            # Ask user for file location to save the text file
            file_path = filedialog.asksaveasfilename(
                defaultextension='.txt',
                filetypes=[("Text files", "*.txt")],
                title="Chọn vị trí lưu file văn bản"
            )

            if file_path:  # If user selects a file path
                with open(file_path, 'w', encoding='utf-8') as file:
                    file.write(email_content)
                messagebox.showinfo("Thành công", "Đã xuất nội dung email ra file văn bản thành công!")
                self.update_status("Đã xuất email thành công")

                # Open the saved text file using default text editor
                os.startfile(file_path)
            else:
                self.update_status("Xuất email bị hủy")

        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi xuất email: {str(e)}")
            self.update_status("Lỗi khi xuất email")

    def format_currency(self, value):
        """Format currency value without decimal places"""
        try:
            if value is None or value == '':
                return ''
            # Convert to float and format with thousand separator, no decimal places
            return f"{float(value):,.0f}"
        except (ValueError, TypeError):
            return str(value)

    def update_readonly_field(self, entry, value):
        """Update readonly field with formatted value"""
        entry.configure(state='normal')
        entry.delete(0, tk.END)
        entry.insert(0, self.format_currency(value))
        entry.configure(state='readonly')

