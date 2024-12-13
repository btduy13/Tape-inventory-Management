# bang_keo_in_tab.py
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tab_base import TabBase
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from tkcalendar import DateEntry
from database import BangKeoInOrder

class BangKeoInTab(TabBase):
    def __init__(self, notebook, parent_form):
        super().__init__(parent_form)
        self.tab = ttk.Frame(notebook)
        notebook.add(self.tab, text="Băng keo In")
        self.db_session = parent_form.db_session

        # Create main frame with padding
        main_frame = ttk.Frame(self.tab, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Configure the grid to expand properly
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=1)
        main_frame.rowconfigure(3, weight=1)
        main_frame.rowconfigure(4, weight=0)  # Buttons row should not expand

        # Build the UI components
        self.build_ui(main_frame)
        self.bind_events()
        self.bind_currency_format()
        self.bind_shortcuts()

    def build_ui(self, main_frame):
        # Configure grid columns
        main_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.columnconfigure(2, weight=1)
        main_frame.columnconfigure(3, weight=1)

        # Title
        title_label = ttk.Label(main_frame, text="BĂNG KEO IN", font=('Segoe UI', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=4, pady=(0, 20), sticky='ew')

        # Basic Information Frame
        basic_info_frame = ttk.LabelFrame(main_frame, text="Thông tin cơ bản", padding=10)
        basic_info_frame.grid(row=1, column=0, columnspan=4, sticky='nsew', padx=5, pady=5)
        self._configure_grid(basic_info_frame, 4)

        # Thêm ID đơn hàng
        ttk.Label(basic_info_frame, text="ID đơn hàng:").grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.id_don_hang = ttk.Entry(basic_info_frame, width=20, state='readonly')
        self.id_don_hang.grid(row=0, column=1, sticky='w', padx=5, pady=5)

        # Tên Hàng
        ttk.Label(basic_info_frame, text="Tên Hàng:").grid(row=0, column=2, sticky='e', padx=5, pady=5)
        self.ten_hang = ttk.Entry(basic_info_frame, width=40)
        self.ten_hang.grid(row=0, column=3, sticky='w', padx=5, pady=5)

        # Ngày dự kiến
        ttk.Label(basic_info_frame, text="Ngày dự kiến giao:").grid(row=1, column=0, sticky='e', padx=5, pady=5)
        self.ngay_du_kien = DateEntry(basic_info_frame, width=15, background='darkblue',
                                    foreground='white', borderwidth=2, date_pattern='dd-mm-yyyy',
                                    locale='vi_VN')
        self.ngay_du_kien.grid(row=1, column=1, sticky='w', padx=5, pady=5)

        # Quy cách
        ttk.Label(basic_info_frame, text="Quy cách:").grid(row=2, column=0, sticky='e', padx=5, pady=5)
        quy_cach_frame = ttk.Frame(basic_info_frame)
        quy_cach_frame.grid(row=2, column=1, columnspan=3, sticky='w', padx=5, pady=5)
        
        self.quy_cach_mm = ttk.Entry(quy_cach_frame, width=10)
        self.quy_cach_mm.pack(side=tk.LEFT, padx=2)
        ttk.Label(quy_cach_frame, text="mm x").pack(side=tk.LEFT)
        self.quy_cach_m = ttk.Entry(quy_cach_frame, width=10)
        self.quy_cach_m.pack(side=tk.LEFT, padx=2)
        ttk.Label(quy_cach_frame, text="m x").pack(side=tk.LEFT)
        self.quy_cach_mic = ttk.Entry(quy_cach_frame, width=10)
        self.quy_cach_mic.pack(side=tk.LEFT, padx=2)
        ttk.Label(quy_cach_frame, text="mic").pack(side=tk.LEFT)

        # Cuộn/1 cây
        ttk.Label(basic_info_frame, text="Cuộn/1 cây:").grid(row=3, column=0, sticky='e', padx=5, pady=5)
        self.cuon_cay = ttk.Entry(basic_info_frame, width=20)
        self.cuon_cay.grid(row=3, column=1, sticky='w', padx=5, pady=5)

        # Chi phí Frame
        cost_frame = ttk.LabelFrame(main_frame, text="Chi phí", padding=10)
        cost_frame.grid(row=2, column=0, columnspan=4, sticky='nsew', padx=5, pady=5)
        self._configure_grid(cost_frame, 4)

        # Số Lượng và Phí số lượng
        ttk.Label(cost_frame, text="Số Lượng:").grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.so_luong = ttk.Entry(cost_frame, width=20)
        self.so_luong.grid(row=0, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(cost_frame, text="Phí số lượng:").grid(row=0, column=2, sticky='e', padx=5, pady=5)
        self.phi_sl = ttk.Entry(cost_frame, width=20)
        self.phi_sl.grid(row=0, column=3, sticky='w', padx=5, pady=5)

        # Màu Keo và Phí Keo
        ttk.Label(cost_frame, text="Màu Keo:").grid(row=1, column=0, sticky='e', padx=5, pady=5)
        self.mau_keo = ttk.Entry(cost_frame, width=20)
        self.mau_keo.grid(row=1, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(cost_frame, text="Phí Keo:").grid(row=1, column=2, sticky='e', padx=5, pady=5)
        self.phi_keo = ttk.Entry(cost_frame, width=20)
        self.phi_keo.grid(row=1, column=3, sticky='w', padx=5, pady=5)

        # Màu Sắc và Phí màu
        ttk.Label(cost_frame, text="Màu Sắc:").grid(row=2, column=0, sticky='e', padx=5, pady=5)
        self.mau_sac = ttk.Entry(cost_frame, width=20)
        self.mau_sac.grid(row=2, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(cost_frame, text="Phí màu:").grid(row=2, column=2, sticky='e', padx=5, pady=5)
        self.phi_mau = ttk.Entry(cost_frame, width=20)
        self.phi_mau.grid(row=2, column=3, sticky='w', padx=5, pady=5)

        # Phí size và Phí cắt
        ttk.Label(cost_frame, text="Phí size:").grid(row=3, column=0, sticky='e', padx=5, pady=5)
        self.phi_size = ttk.Entry(cost_frame, width=20)
        self.phi_size.grid(row=3, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(cost_frame, text="Phí cắt:").grid(row=3, column=2, sticky='e', padx=5, pady=5)
        self.phi_cat = ttk.Entry(cost_frame, width=20)
        self.phi_cat.grid(row=3, column=3, sticky='w', padx=5, pady=5)

        # Giá cả Frame
        price_frame = ttk.LabelFrame(main_frame, text="Giá cả", padding=10)
        price_frame.grid(row=3, column=0, columnspan=4, sticky='nsew', padx=5, pady=5)
        self._configure_grid(price_frame, 4)

        # Đơn giá vốn
        ttk.Label(price_frame, text="Đơn giá vốn:").grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.don_gia_von = ttk.Entry(price_frame, width=20)
        self.don_gia_von.grid(row=0, column=1, sticky='w', padx=5, pady=5)

        # Đơn giá gốc
        ttk.Label(price_frame, text="Đơn giá gốc:").grid(row=1, column=0, sticky='e', padx=5, pady=5)
        self.don_gia_goc = ttk.Entry(price_frame, width=20, state='readonly')
        self.don_gia_goc.grid(row=1, column=1, sticky='w', padx=5, pady=5)

        # Thành Tiền gốc
        ttk.Label(price_frame, text="Thành Tiền(gốc):").grid(row=2, column=0, sticky='e', padx=5, pady=5)
        self.thanh_tien_goc = ttk.Entry(price_frame, width=20, state='readonly')
        self.thanh_tien_goc.grid(row=2, column=1, sticky='w', padx=5, pady=5)

        # Đơn giá bán
        ttk.Label(price_frame, text="Đơn giá(bán):").grid(row=0, column=2, sticky='e', padx=5, pady=5)
        self.don_gia_ban = ttk.Entry(price_frame, width=20)
        self.don_gia_ban.grid(row=0, column=3, sticky='w', padx=5, pady=5)

        # Thành Tiền bán
        ttk.Label(price_frame, text="Thành Tiền(bán):").grid(row=1, column=2, sticky='e', padx=5, pady=5)
        self.thanh_tien_ban = ttk.Entry(price_frame, width=20, state='readonly')
        self.thanh_tien_ban.grid(row=1, column=3, sticky='w', padx=5, pady=5)

        # Tiền cọc
        ttk.Label(price_frame, text="Tiền cọc:").grid(row=2, column=2, sticky='e', padx=5, pady=5)
        self.tien_coc = ttk.Entry(price_frame, width=20)
        self.tien_coc.grid(row=2, column=3, sticky='w', padx=5, pady=5)

        # Công nợ khách
        ttk.Label(price_frame, text="Công nợ khách:").grid(row=3, column=0, sticky='e', padx=5, pady=5)
        self.cong_no_khach = ttk.Entry(price_frame, width=20, state='readonly')
        self.cong_no_khach.grid(row=3, column=1, sticky='w', padx=5, pady=5)

        # CTV và Hoa hồng
        ttk.Label(price_frame, text="CTV:").grid(row=4, column=0, sticky='e', padx=5, pady=5)
        self.ctv = ttk.Entry(price_frame, width=20)
        self.ctv.grid(row=4, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(price_frame, text="Hoa hồng(%):").grid(row=4, column=2, sticky='e', padx=5, pady=5)
        self.hoa_hong = ttk.Entry(price_frame, width=20)
        self.hoa_hong.grid(row=4, column=3, sticky='w', padx=5, pady=5)

        # Tiền hoa hồng
        ttk.Label(price_frame, text="Tiền hoa hồng:").grid(row=5, column=0, sticky='e', padx=5, pady=5)
        self.tien_hoa_hong = ttk.Entry(price_frame, width=20, state='readonly')
        self.tien_hoa_hong.grid(row=5, column=1, sticky='w', padx=5, pady=5)

        # Lõi Giấy và Thùng/Bao
        ttk.Label(price_frame, text="Lõi Giấy:").grid(row=6, column=0, sticky='e', padx=5, pady=5)
        self.loi_giay = ttk.Entry(price_frame, width=20)
        self.loi_giay.grid(row=6, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(price_frame, text="Thùng/Bao:").grid(row=6, column=2, sticky='e', padx=5, pady=5)
        self.thung_bao = ttk.Entry(price_frame, width=20)
        self.thung_bao.grid(row=6, column=3, sticky='w', padx=5, pady=5)

        # Lợi nhuận
        ttk.Label(price_frame, text="Lợi nhuận:").grid(row=7, column=0, sticky='e', padx=5, pady=5)
        self.loi_nhuan = ttk.Entry(price_frame, width=20, state='readonly')
        self.loi_nhuan.grid(row=7, column=1, sticky='w', padx=5, pady=5)

        # Buttons Frame
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, columnspan=4, pady=20, sticky='e')  # Align to the right

        # Define a consistent style for all buttons
        style = ttk.Style()
        style.configure('CustomButton.TButton',
                        font=('Segoe UI', 10),
                        padding=6)

        # Create buttons with the custom style
        btn_tinh_toan = ttk.Button(button_frame, text="Tính toán", command=self.tinh_toan, style='CustomButton.TButton', width=12)
        btn_luu = ttk.Button(button_frame, text="Lưu", command=self.luu_don_hang, style='CustomButton.TButton', width=12)
        btn_xuat_excel = ttk.Button(button_frame, text="Xuất Excel", command=self.export_to_excel, style='CustomButton.TButton', width=12)
        btn_xuat_email = ttk.Button(button_frame, text="Xuất Email", command=self.export_email, style='CustomButton.TButton', width=12)
        btn_xoa = ttk.Button(button_frame, text="Xóa", command=self.xoa_form, style='CustomButton.TButton', width=12)
        btn_thoat = ttk.Button(button_frame, text="Thoát", command=self.thoat, style='CustomButton.TButton', width=12)

        # Pack buttons to the right with consistent padding
        btn_thoat.pack(side='right', padx=5)
        btn_xoa.pack(side='right', padx=5)
        btn_xuat_email.pack(side='right', padx=5)
        btn_xuat_excel.pack(side='right', padx=5)
        btn_luu.pack(side='right', padx=5)
        btn_tinh_toan.pack(side='right', padx=5)

        # Set focus to the first entry
        self.ten_hang.focus_set()

    def _configure_grid(self, frame, cols):
        """Configure grid columns to be evenly spaced"""
        for i in range(cols):
            frame.columnconfigure(i, weight=1)

    def bind_events(self):
        entries_to_bind = [
            self.so_luong, self.phi_sl, self.phi_keo, self.phi_size,
            self.phi_cat, self.don_gia_von, self.don_gia_ban, self.tien_coc,
            self.hoa_hong, self.phi_mau, self.quy_cach_mm,
            self.quy_cach_m, self.quy_cach_mic, self.cuon_cay
        ]
        for entry in entries_to_bind:
            entry.bind('<KeyRelease>', self.auto_calculate)

        # Register validation command
        vcmd = (self.root.register(self.is_valid_float), '%P')
        entries_to_validate = entries_to_bind
        for entry in entries_to_validate:
            entry.config(validate='key', validatecommand=vcmd)

    def auto_calculate(self, event):
        self.tinh_toan()

    def bind_currency_format(self):
        currency_fields = [
            self.don_gia_von,
            self.don_gia_ban,
            self.tien_coc,
            self.phi_sl,
            self.phi_keo,
            self.phi_mau,
            self.phi_size,
            self.phi_cat,
        ]
        for field in currency_fields:
            field.bind('<FocusOut>', self.format_currency_input)

    def bind_shortcuts(self):
        self.root.bind('<Control-s>', lambda event: self.luu_don_hang())
        self.root.bind('<Control-t>', lambda event: self.tinh_toan())
        self.root.bind('<Control-e>', lambda event: self.export_to_excel())
        self.root.bind('<Control-q>', lambda event: self.thoat())

    def tinh_toan(self):
        try:
            # 1. Các giá trị đầu vào cơ bản
            don_gia_von = self.validate_float_input(self.don_gia_von.get())
            phi_sl = self.validate_float_input(self.phi_sl.get())
            phi_mau = self.validate_float_input(self.phi_mau.get())
            phi_keo = self.validate_float_input(self.phi_keo.get())
            phi_size = self.validate_float_input(self.phi_size.get())
            phi_cat = self.validate_float_input(self.phi_cat.get())
            so_luong = self.validate_float_input(self.so_luong.get())
            don_gia_ban = self.validate_float_input(self.don_gia_ban.get())
            tien_coc = self.validate_float_input(self.tien_coc.get())
            hoa_hong = self.validate_float_input(self.hoa_hong.get()) / 100  # Chuyển % thành decimal

            # 2. Tính đơn giá gốc
            cuon_cay = self.validate_float_input(self.cuon_cay.get())
            quy_cach_m = self.validate_float_input(self.quy_cach_m.get())

            if cuon_cay == 0 or quy_cach_m == 0:
                don_gia_goc = 0
            else:
                # Công thức tính đơn giá gốc:
                # (Tổng chi phí) / 90 * quy_cach_m / cuon_cay
                don_gia_goc = (don_gia_von + phi_sl + phi_mau + phi_keo + phi_size + phi_cat) / 90 * quy_cach_m / cuon_cay

            # 3. Tính các giá trị phái sinh
            thanh_tien_goc = don_gia_goc * so_luong  # Thành tiền gốc = đơn giá gốc × số lượng
            thanh_tien_ban = don_gia_ban * so_luong  # Thành tiền bán = đơn giá bán × số lượng
            cong_no_khach = thanh_tien_ban - tien_coc  # Công nợ = thành tiền bán - tiền cọc
            loi_nhuan = thanh_tien_ban - thanh_tien_goc  # Lợi nhuận = thành tiền bán - thành tiền gốc
            tien_hoa_hong = loi_nhuan * hoa_hong  # Tiền hoa hồng = lợi nhuận × tỷ lệ hoa hồng

            # Cập nhật giao diện
            self.update_readonly_field(self.don_gia_goc, don_gia_goc)
            self.update_readonly_field(self.thanh_tien_goc, thanh_tien_goc)
            self.update_readonly_field(self.thanh_tien_ban, thanh_tien_ban)
            self.update_readonly_field(self.cong_no_khach, cong_no_khach)
            self.update_readonly_field(self.tien_hoa_hong, tien_hoa_hong)
            self.update_readonly_field(self.loi_nhuan, loi_nhuan)

            # Update status bar
            self.update_status("Tính toán thành công")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi tính toán: {str(e)}")
            self.update_status("Lỗi khi tính toán")

    def luu_don_hang(self):
        try:
            # Kiểm tra các trường bắt buộc
            required_fields = {
                'Tên hàng': self.ten_hang.get(),
                'Số lượng': self.so_luong.get(),
                'Đơn giá bán': self.don_gia_ban.get()
            }
            
            empty_fields = [field for field, value in required_fields.items() if not value.strip()]
            if empty_fields:
                messagebox.showwarning("Cảnh báo", f"Vui lòng điền đầy đủ thông tin: {', '.join(empty_fields)}")
                return
            
            # Lấy dữ liệu từ form
            data = {
                'thoi_gian': datetime.now(),
                'ten_hang': self.ten_hang.get(),
                'ngay_du_kien': self.ngay_du_kien.get_date(),
                'quy_cach_mm': self.parse_float(self.quy_cach_mm.get()),
                'quy_cach_m': self.parse_float(self.quy_cach_m.get()),
                'quy_cach_mic': self.parse_float(self.quy_cach_mic.get()),
                'cuon_cay': self.parse_float(self.cuon_cay.get()),
                'so_luong': self.parse_float(self.so_luong.get()),
                'phi_sl': self.parse_float(self.phi_sl.get()),
                'mau_keo': self.mau_keo.get(),
                'phi_keo': self.parse_float(self.phi_keo.get()),
                'mau_sac': self.mau_sac.get(),
                'phi_mau': self.parse_float(self.phi_mau.get()),
                'phi_size': self.parse_float(self.phi_size.get()),
                'phi_cat': self.parse_float(self.phi_cat.get()),
                'don_gia_von': self.parse_float(self.don_gia_von.get()),
                'don_gia_goc': self.parse_float(self.don_gia_goc.get()),
                'thanh_tien_goc': self.parse_float(self.thanh_tien_goc.get()),
                'don_gia_ban': self.parse_float(self.don_gia_ban.get()),
                'thanh_tien_ban': self.parse_float(self.thanh_tien_ban.get()),
                'tien_coc': self.parse_float(self.tien_coc.get()),
                'cong_no_khach': self.parse_float(self.cong_no_khach.get()),
                'ctv': self.ctv.get(),
                'hoa_hong': self.parse_float(self.hoa_hong.get()),
                'tien_hoa_hong': self.parse_float(self.tien_hoa_hong.get()),
                'loi_giay': self.loi_giay.get(),
                'thung_bao': self.thung_bao.get(),
                'loi_nhuan': self.parse_float(self.loi_nhuan.get()),
                'da_giao': self.da_giao,
                'da_tat_toan': self.da_tat_toan
            }
            
            # Tạo đối tượng BangKeoInOrder mới
            don_hang = BangKeoInOrder(**data)
            
            # Thêm vào database
            self.db_session.add(don_hang)
            self.db_session.commit()
            
            # Cập nhật ID đơn hàng trên form
            self.id_don_hang.configure(state='normal')
            self.id_don_hang.delete(0, tk.END)
            self.id_don_hang.insert(0, str(don_hang.id))
            self.id_don_hang.configure(state='readonly')
            
            messagebox.showinfo("Thành công", "Đã lưu đơn hàng thành công!")
            
            # Cập nhật history tab và thống kê tab
            if hasattr(self.parent_form, 'history_tab'):
                self.parent_form.history_tab.refresh_data()
            if hasattr(self.parent_form, 'thong_ke_tab'):
                self.parent_form.thong_ke_tab.load_data()
                
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lưu đơn hàng: {str(e)}")
            self.db_session.rollback()

    def xoa_form(self):
        try:
            if messagebox.askyesno("Xác nhận", "Bạn có chắc muốn xóa toàn bộ form?"):
                # Xóa nội dung của tất cả các Entry trong tab này
                for child in self.tab.winfo_children():
                    self.clear_widget(child)
                
                # Xóa ID đơn hàng
                self.id_don_hang.configure(state='normal')
                self.id_don_hang.delete(0, tk.END)
                self.id_don_hang.configure(state='readonly')
                
                messagebox.showinfo("Thành công", "Đã xóa form")
                self.update_status("Form đã được xóa")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi xóa form: {str(e)}")
            self.update_status("Lỗi khi xóa form")

    def thoat(self):
        try:
            if messagebox.askyesno("Xác nhận", "Bạn có chắc muốn thoát chương trình?"):
                self.root.quit()
                self.update_status("Thoát chương trình")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi thoát: {str(e)}")
            self.update_status("Lỗi khi thoát chương trình")

    def export_to_excel(self):
        try:
            # Chọn vị trí lưu file
            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[("Excel files", "*.xlsx")],
                title="Chọn vị trí lưu file Excel"
            )

            if not file_path:  # Nếu người dùng hủy
                return

            # Tạo dữ liệu với ngày là trường đầu tiên
            current_date = datetime.now().strftime('%d-%m-%Y')

            data = {
                'ID': self.id_don_hang.get(),
                'Ngày': current_date,
                'Tên Hàng': self.ten_hang.get(),
                'Ngày dự kiến': self.ngay_du_kien.get_date().strftime('%d-%m-%Y'),
                'Quy Cách (mm)': self.quy_cach_mm.get(),
                'Quy Cách (m)': self.quy_cach_m.get(),
                'Quy Cách (mic)': self.quy_cach_mic.get(),
                'Cuộn/1 cây': self.cuon_cay.get(),
                'Số Lượng': self.so_luong.get(),
                'Phí số lượng': self.phi_sl.get(),
                'Màu Keo': self.mau_keo.get(),
                'Phí Keo': self.phi_keo.get(),
                'Màu Sắc': self.mau_sac.get(),
                'Phí màu': self.phi_mau.get(),
                'Phí size': self.phi_size.get(),
                'Phí cắt': self.phi_cat.get(),
                'Đơn giá vốn': self.don_gia_von.get(),
                'Đơn giá gốc': self.don_gia_goc.get(),
                'Thành Tiền(gốc)': self.thanh_tien_goc.get(),
                'Đơn giá(bán)': self.don_gia_ban.get(),
                'Thành Tiền(bán)': self.thanh_tien_ban.get(),
                'Tiền cọc': self.tien_coc.get(),
                'Công nợ khách': self.cong_no_khach.get(),
                'CTV': self.ctv.get(),
                'Hoa hồng(%)': self.hoa_hong.get(),
                'Tiền hoa hồng': self.tien_hoa_hong.get(),
                'Lõi Giấy': self.loi_giay.get(),
                'Thùng/Bao': self.thung_bao.get(),
                'Lợi nhuận': self.loi_nhuan.get()
            }

            # Kiểm tra file có tồn tại không
            if os.path.exists(file_path):
                # Nếu file đã tồn tại, mở file và thêm dữ liệu mới
                wb = load_workbook(file_path)
                if "Bang keo in" in wb.sheetnames:
                    ws = wb["Bang keo in"]
                    next_row = ws.max_row + 1
                else:
                    ws = wb.create_sheet("Bang keo in")
                    # Ghi headers cho sheet mới
                    for col, header in enumerate(data.keys(), 1):
                        ws.cell(row=1, column=col, value=header)
                    next_row = 2
            else:
                # Tạo file mới nếu chưa tồn tại
                wb = Workbook()
                ws = wb.active
                ws.title = "Bang keo in"
                # Ghi headers
                for col, header in enumerate(data.keys(), 1):
                    ws.cell(row=1, column=col, value=header)
                next_row = 2

            # Ghi data vào dòng tiếp theo
            for col, value in enumerate(data.values(), 1):
                ws.cell(row=next_row, column=col, value=value)

            # Lưu file
            wb.save(file_path)

            messagebox.showinfo("Thành công", "Đã xuất dữ liệu ra Excel thành công!")
            self.update_status("Đã xuất Excel thành công")

        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi xuất Excel: {str(e)}")
            self.update_status("Lỗi khi xuất Excel")

    def export_email(self):
        """Export the order details to a text file and open it."""
        ten_hang = self.ten_hang.get()
        mau_sac = self.mau_sac.get()
        mau_keo = self.mau_keo.get()
        so_luong = self.so_luong.get()
        quy_cach = f"{self.quy_cach_mm.get()}mm * {self.quy_cach_m.get()}m * {self.quy_cach_mic.get()}mic"
        loi_giay = self.loi_giay.get()
        thung_bao = self.thung_bao.get()

        email_content = (
            f"Chào bác,\n\n"
            f"Bác làm giúp con đơn hàng in logo \"{ten_hang}\" này nhé\n"
            f"Màu sắc: {mau_sac} / Màu keo: {mau_keo}\n"
            f"Số lượng: {so_luong} cuộn\n"
            f"Quy cách: {quy_cach}\n"
            f"Lõi giấy: {loi_giay} - Thùng bao: {thung_bao}\n\n"
            f"Cám ơn bác\n"
            f"Quế"
        )

        # Ask user for file location to save the text file
        file_path = filedialog.asksaveasfilename(
            defaultextension='.txt',
            filetypes=[("Text files", "*.txt")],
            title="Chọn vị trí lưu file văn bản"
        )

        if file_path:  # If user selects a file path
            with open(file_path, 'w', encoding='utf-8') as file:
                file.write(email_content)
            messagebox.showinfo("Thành công", "Đã xuất nội dung email ra file văn bản thành công!")
            self.update_status("Đã xuất email thành công")

            # Open the saved text file using default text editor
            os.startfile(file_path)
        else:
            self.update_status("Xuất email bị hủy")
