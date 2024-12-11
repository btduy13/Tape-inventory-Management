# truc_in_tab.py
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tab_base import TabBase
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from tkcalendar import DateEntry
from database import TrucInOrder

class TrucInTab(TabBase):
    def __init__(self, notebook, parent_form):
        super().__init__(parent_form)
        self.tab = ttk.Frame(notebook)
        notebook.add(self.tab, text="Trục in")
        self.db_session = parent_form.db_session

        # Create main frame with padding
        main_frame = ttk.Frame(self.tab, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Build the UI components
        self.build_ui(main_frame)
        self.bind_events()
        self.bind_currency_format()
        self.bind_shortcuts()
        
        # Checkbox cho trạng thái
        self.da_giao_var = tk.BooleanVar(value=False)
        self.da_tat_toan_var = tk.BooleanVar(value=False)

    def build_ui(self, main_frame):
        # Configure grid columns
        main_frame.columnconfigure(1, weight=1)
        main_frame.columnconfigure(3, weight=1)

        # Title
        title_label = ttk.Label(main_frame, text="TRỤC IN", font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=4, pady=(0, 20))

        # Basic Information Frame
        basic_info_frame = ttk.LabelFrame(main_frame, text="Thông tin trục in", padding=10)
        basic_info_frame.grid(row=1, column=0, columnspan=4, sticky='nsew', padx=5, pady=5)
        self._configure_grid(basic_info_frame, 4)

        # Thêm ID đơn hàng
        ttk.Label(basic_info_frame, text="ID đơn hàng:").grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.id_don_hang = ttk.Entry(basic_info_frame, width=20, state='readonly')
        self.id_don_hang.grid(row=0, column=1, sticky='w', padx=5, pady=5)

        # Tên Hàng
        ttk.Label(basic_info_frame, text="Tên Hàng:").grid(row=1, column=0, sticky='e', padx=5, pady=5)
        self.truc_in_ten_hang = ttk.Entry(basic_info_frame, width=40)
        self.truc_in_ten_hang.grid(row=1, column=1, sticky='w', padx=5, pady=5)

        # Ngày dự kiến
        ttk.Label(basic_info_frame, text="Ngày dự kiến giao:").grid(row=1, column=2, sticky='e', padx=5, pady=5)
        self.truc_in_ngay_du_kien = DateEntry(basic_info_frame, width=15, background='darkblue',
                                            foreground='white', borderwidth=2, date_pattern='dd-mm-yyyy',
                                            locale='vi_VN')
        self.truc_in_ngay_du_kien.grid(row=1, column=3, sticky='w', padx=5, pady=5)

        # Quy cách
        ttk.Label(basic_info_frame, text="Quy cách:").grid(row=2, column=0, sticky='e', padx=5, pady=5)
        quy_cach_frame = ttk.Frame(basic_info_frame)
        quy_cach_frame.grid(row=2, column=1, sticky='w', padx=5, pady=5)
        
        self.truc_in_quy_cach = ttk.Entry(quy_cach_frame, width=20)
        self.truc_in_quy_cach.pack(side=tk.LEFT, padx=(0,2))
        ttk.Label(quy_cach_frame, text="mm").pack(side=tk.LEFT)

        # Số lượng
        ttk.Label(basic_info_frame, text="Số lượng:").grid(row=3, column=0, sticky='e', padx=5, pady=5)
        self.truc_in_so_luong = ttk.Entry(basic_info_frame, width=20)
        self.truc_in_so_luong.grid(row=3, column=1, sticky='w', padx=5, pady=5)

        # Màu sắc
        ttk.Label(basic_info_frame, text="Màu sắc:").grid(row=4, column=0, sticky='e', padx=5, pady=5)
        self.truc_in_mau_sac = ttk.Entry(basic_info_frame, width=20)
        self.truc_in_mau_sac.grid(row=4, column=1, sticky='w', padx=5, pady=5)

        # Màu keo
        ttk.Label(basic_info_frame, text="Màu keo:").grid(row=5, column=0, sticky='e', padx=5, pady=5)
        self.truc_in_mau_keo = ttk.Entry(basic_info_frame, width=20)
        self.truc_in_mau_keo.grid(row=5, column=1, sticky='w', padx=5, pady=5)

        # Giá cả Frame
        price_frame = ttk.LabelFrame(main_frame, text="Giá cả", padding=10)
        price_frame.grid(row=2, column=0, columnspan=4, sticky='nsew', padx=5, pady=5)
        self._configure_grid(price_frame, 4)

        # Đơn giá gốc
        ttk.Label(price_frame, text="Đơn giá gốc:").grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.truc_in_don_gia_goc = ttk.Entry(price_frame, width=20)
        self.truc_in_don_gia_goc.grid(row=0, column=1, sticky='w', padx=5, pady=5)

        # Thành tiền
        ttk.Label(price_frame, text="Thành tiền:").grid(row=1, column=0, sticky='e', padx=5, pady=5)
        self.truc_in_thanh_tien = ttk.Entry(price_frame, width=20, state='readonly')
        self.truc_in_thanh_tien.grid(row=1, column=1, sticky='w', padx=5, pady=5)

        # Đơn giá bán
        ttk.Label(price_frame, text="Đơn giá (bán):").grid(row=0, column=2, sticky='e', padx=5, pady=5)
        self.truc_in_don_gia_ban = ttk.Entry(price_frame, width=20)
        self.truc_in_don_gia_ban.grid(row=0, column=3, sticky='w', padx=5, pady=5)

        # Thành tiền bán
        ttk.Label(price_frame, text="Thành tiền (bán):").grid(row=1, column=2, sticky='e', padx=5, pady=5)
        self.truc_in_thanh_tien_ban = ttk.Entry(price_frame, width=20, state='readonly')
        self.truc_in_thanh_tien_ban.grid(row=1, column=3, sticky='w', padx=5, pady=5)

        # Công nợ khách
        ttk.Label(price_frame, text="Công nợ khách:").grid(row=2, column=0, sticky='e', padx=5, pady=5)
        self.truc_in_cong_no_khach = ttk.Entry(price_frame, width=20, state='readonly')
        self.truc_in_cong_no_khach.grid(row=2, column=1, sticky='w', padx=5, pady=5)

        # CTV và Hoa hồng
        ttk.Label(price_frame, text="CTV:").grid(row=3, column=0, sticky='e', padx=5, pady=5)
        self.truc_in_ctv = ttk.Entry(price_frame, width=20)
        self.truc_in_ctv.grid(row=3, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(price_frame, text="Hoa Hồng (%):").grid(row=3, column=2, sticky='e', padx=5, pady=5)
        self.truc_in_hoa_hong = ttk.Entry(price_frame, width=20)
        self.truc_in_hoa_hong.grid(row=3, column=3, sticky='w', padx=5, pady=5)

        # Tiền hoa hồng
        ttk.Label(price_frame, text="Tiền hoa hồng:").grid(row=4, column=0, sticky='e', padx=5, pady=5)
        self.truc_in_tien_hoa_hong = ttk.Entry(price_frame, width=20, state='readonly')
        self.truc_in_tien_hoa_hong.grid(row=4, column=1, sticky='w', padx=5, pady=5)

        # Lợi nhuận
        ttk.Label(price_frame, text="Lợi nhuận:").grid(row=5, column=0, sticky='e', padx=5, pady=5)
        self.truc_in_loi_nhuan = ttk.Entry(price_frame, width=20, state='readonly')
        self.truc_in_loi_nhuan.grid(row=5, column=1, sticky='w', padx=5, pady=5)

        # Buttons Frame
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=3, column=0, columnspan=4, pady=20)

        self.create_button(button_frame, "Tính toán", self.tinh_toan_truc_in).pack(side=tk.LEFT, padx=5)
        self.create_button(button_frame, "Lưu", self.luu_truc_in).pack(side=tk.LEFT, padx=5)
        self.create_button(button_frame, "Xuất Excel", self.export_to_excel).pack(side=tk.LEFT, padx=5)
        self.create_button(button_frame, "Xuất Email", self.export_truc_in_email).pack(side=tk.LEFT, padx=5)
        self.create_button(button_frame, "Xóa", self.xoa_form_truc_in).pack(side=tk.LEFT, padx=5)

        # Set focus to the first entry
        self.truc_in_ten_hang.focus_set()

    def _configure_grid(self, frame, cols):
        """Configure grid columns to be evenly spaced"""
        for i in range(cols):
            frame.columnconfigure(i, weight=1)

    def bind_events(self):
        entries_to_bind = [
            self.truc_in_so_luong,
            self.truc_in_don_gia_ban,
            self.truc_in_don_gia_goc,
            self.truc_in_hoa_hong,
        ]
        for entry in entries_to_bind:
            entry.bind('<KeyRelease>', self.auto_calculate)

        # Register validation command
        vcmd = (self.root.register(self.is_valid_float), '%P')
        for entry in entries_to_bind:
            entry.config(validate='key', validatecommand=vcmd)

    def auto_calculate(self, event):
        self.tinh_toan_truc_in()

    def bind_currency_format(self):
        currency_fields = [
            self.truc_in_don_gia_goc,
            self.truc_in_don_gia_ban,
        ]
        for field in currency_fields:
            field.bind('<FocusOut>', self.format_currency_input)

    def bind_shortcuts(self):
        self.root.bind('<Control-s>', lambda event: self.luu_truc_in())
        self.root.bind('<Control-t>', lambda event: self.tinh_toan_truc_in())
        self.root.bind('<Control-e>', lambda event: self.export_to_excel())
        self.root.bind('<Control-q>', lambda event: self.root.quit())

    def tinh_toan_truc_in(self):
        try:
            # Get values
            so_luong = self.validate_float_input(self.truc_in_so_luong.get())
            don_gia_ban = self.validate_float_input(self.truc_in_don_gia_ban.get())
            don_gia_goc = self.validate_float_input(self.truc_in_don_gia_goc.get())
            hoa_hong = self.validate_float_input(self.truc_in_hoa_hong.get()) / 100

            # Calculate values
            thanh_tien = don_gia_goc * so_luong  # Thành tiền gốc
            thanh_tien_ban = don_gia_ban * so_luong  # Thành tiền bán
            loi_nhuan = thanh_tien_ban - thanh_tien  # Lợi nhuận
            tien_hoa_hong = loi_nhuan * hoa_hong  # Tiền hoa hồng
            cong_no_khach = thanh_tien_ban  # Công nợ khách = thành tiền bán

            # Update fields
            self.update_readonly_field(self.truc_in_thanh_tien, thanh_tien)
            self.update_readonly_field(self.truc_in_thanh_tien_ban, thanh_tien_ban)
            self.update_readonly_field(self.truc_in_cong_no_khach, cong_no_khach)
            self.update_readonly_field(self.truc_in_tien_hoa_hong, tien_hoa_hong)
            self.update_readonly_field(self.truc_in_loi_nhuan, loi_nhuan)

            self.update_status("Tính toán trục in thành công")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi tính toán trục in: {str(e)}")
            self.update_status("Lỗi khi tính toán trục in")

    def luu_truc_in(self):
        try:
            # Kiểm tra các trường bắt buộc
            required_fields = {
                'Tên hàng': self.truc_in_ten_hang.get(),
                'Số lượng': self.truc_in_so_luong.get(),
                'Đơn giá bán': self.truc_in_don_gia_ban.get()
            }
            
            empty_fields = [field for field, value in required_fields.items() if not value.strip()]
            if empty_fields:
                messagebox.showwarning("Cảnh báo", f"Vui lòng điền đầy đủ thông tin: {', '.join(empty_fields)}")
                return
            
            # Lấy dữ liệu từ form
            data = {
                'thoi_gian': datetime.now(),
                'ten_hang': self.truc_in_ten_hang.get(),
                'ngay_du_kien': self.truc_in_ngay_du_kien.get_date(),
                'quy_cach': self.truc_in_quy_cach.get(),
                'so_luong': self.parse_float(self.truc_in_so_luong.get()),
                'mau_sac': self.truc_in_mau_sac.get(),
                'mau_keo': self.truc_in_mau_keo.get(),
                'don_gia_goc': self.parse_float(self.truc_in_don_gia_goc.get()),
                'thanh_tien': self.parse_float(self.truc_in_thanh_tien.get()),
                'don_gia_ban': self.parse_float(self.truc_in_don_gia_ban.get()),
                'thanh_tien_ban': self.parse_float(self.truc_in_thanh_tien_ban.get()),
                'cong_no_khach': self.parse_float(self.truc_in_cong_no_khach.get()),
                'ctv': self.truc_in_ctv.get(),
                'hoa_hong': self.parse_float(self.truc_in_hoa_hong.get()),
                'tien_hoa_hong': self.parse_float(self.truc_in_tien_hoa_hong.get()),
                'loi_nhuan': self.parse_float(self.truc_in_loi_nhuan.get()),
                'da_giao': self.da_giao_var.get(),
                'da_tat_toan': self.da_tat_toan_var.get()
            }
            
            # Tạo đối tượng TrucInOrder mới
            don_hang = TrucInOrder(**data)
            
            # Thêm vào database
            self.db_session.add(don_hang)
            self.db_session.commit()
            
            # Cập nhật ID đơn hàng trên form
            self.id_don_hang.configure(state='normal')
            self.id_don_hang.delete(0, tk.END)
            self.id_don_hang.insert(0, str(don_hang.id))
            self.id_don_hang.configure(state='readonly')
            
            messagebox.showinfo("Thành công", "Đã lưu đơn hàng thành công!")
            
            # Cập nhật history tab và thống kê tab
            if hasattr(self.parent_form, 'history_tab'):
                self.parent_form.history_tab.refresh_data()
            if hasattr(self.parent_form, 'thong_ke_tab'):
                self.parent_form.thong_ke_tab.load_data()
                
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lưu đơn hàng: {str(e)}")
            self.db_session.rollback()

    def export_to_excel(self):
        try:
            # Chọn vị trí lưu file
            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[("Excel files", "*.xlsx")],
                title="Chọn vị trí lưu file Excel"
            )

            if not file_path:  # Nếu người dùng hủy
                return

            # Tạo dữ liệu với ngày là trường đầu tiên
            current_date = datetime.now().strftime('%d-%m-%Y')

            data = {
                'ID': self.id_don_hang.get(),
                'Ngày': current_date,
                'Tên Hàng': self.truc_in_ten_hang.get(),
                'Ngày dự kiến': self.truc_in_ngay_du_kien.get_date().strftime('%d-%m-%Y'),
                'Quy Cách': f"{self.truc_in_quy_cach.get()} mm",
                'Số Lượng': self.truc_in_so_luong.get(),
                'Màu Sắc': self.truc_in_mau_sac.get(),
                'Màu Keo': self.truc_in_mau_keo.get(),
                'Đơn giá gốc': self.truc_in_don_gia_goc.get(),
                'Thành tiền': self.truc_in_thanh_tien.get(),
                'Đơn giá bán': self.truc_in_don_gia_ban.get(),
                'Thành tiền bán': self.truc_in_thanh_tien_ban.get(),
                'Công nợ khách': self.truc_in_cong_no_khach.get(),
                'CTV': self.truc_in_ctv.get(),
                'Hoa Hồng(%)': self.truc_in_hoa_hong.get(),
                'Tiền hoa hồng': self.truc_in_tien_hoa_hong.get(),
                'Lợi nhuận': self.truc_in_loi_nhuan.get()
            }

            # Tạo hoặc tải workbook
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                ws = wb.active
                next_row = ws.max_row + 1
            else:
                wb = Workbook()
                ws = wb.active
                # Thêm headers
                for col, header in enumerate(data.keys(), 1):
                    ws.cell(row=1, column=col, value=header)
                next_row = 2

            # Thêm dữ liệu
            for col, value in enumerate(data.values(), 1):
                ws.cell(row=next_row, column=col, value=value)

            # Lưu workbook
            wb.save(file_path)

            messagebox.showinfo("Thành công", "Đã xuất dữ liệu ra Excel thành công!")
            self.update_status("Đã xuất Excel thành công")

        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi xuất Excel: {str(e)}")
            self.update_status("Lỗi khi xuất Excel")

    def export_truc_in_email(self):
        try:
            ten_hang = self.truc_in_ten_hang.get()
            mau_sac = self.truc_in_mau_sac.get()
            mau_keo = self.truc_in_mau_keo.get()
            so_luong = self.truc_in_so_luong.get()
            quy_cach = f"{self.truc_in_quy_cach.get()}mm"

            email_content = (
                f"Chào bác,\n\n"
                f"Bác làm giúp con đơn hàng trc in \"{ten_hang}\" này nhé\n"
                f"Màu sắc: {mau_sac} / Màu keo: {mau_keo}\n"
                f"Số lượng: {so_luong} cái\n"
                f"Quy cách: {quy_cach}\n\n"
                f"Cám ơn bác\n"
                f"Quế"
            )

            file_path = filedialog.asksaveasfilename(
                defaultextension='.txt',
                filetypes=[("Text files", "*.txt")],
                title="Chọn vị trí lưu file văn bản"
            )

            if file_path:
                with open(file_path, 'w', encoding='utf-8') as file:
                    file.write(email_content)
                messagebox.showinfo("Thành công", "Đã xuất nội dung email trục in ra file văn bản thành công!")
                self.update_status("Đã xuất email trục in thành công")

                # Open the file after saving
                os.startfile(file_path)
            else:
                self.update_status("Xuấất email trục in bị hủy")

        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi xuất email trục in: {str(e)}")
            self.update_status("Lỗi khi xuất email trục in")

    def xoa_form_truc_in(self):
        try:
            if messagebox.askyesno("Xác nhận", "Bạn có chắc muốn xóa toàn bộ form trục in?"):
                fields = [
                    self.truc_in_ten_hang, self.truc_in_quy_cach, self.truc_in_so_luong,
                    self.truc_in_mau_sac, self.truc_in_mau_keo, self.truc_in_don_gia_ban,
                    self.truc_in_ctv, self.truc_in_hoa_hong
                ]
                for field in fields:
                    field.delete(0, tk.END)

                # Clear readonly fields
                readonly_fields = [
                    self.truc_in_don_gia_goc, self.truc_in_thanh_tien,
                    self.truc_in_thanh_tien_ban, self.truc_in_cong_no_khach,
                    self.truc_in_tien_hoa_hong, self.truc_in_loi_nhuan
                ]
                for field in readonly_fields:
                    field.configure(state='normal')
                    field.delete(0, tk.END)
                    field.configure(state='readonly')
                    
                # Xóa ID đơn hàng
                self.id_don_hang.configure(state='normal')
                self.id_don_hang.delete(0, tk.END)
                self.id_don_hang.configure(state='readonly')

                self.update_status("Đã xóa form trục in")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi xóa form trục in: {str(e)}")
            self.update_status("Lỗi khi xóa form trục in")
