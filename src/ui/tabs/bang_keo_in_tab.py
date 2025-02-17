# bang_keo_in_tab.py
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from src.ui.tabs.tab_base import TabBase
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from tkcalendar import DateEntry
from src.database.database import BangKeoInOrder

class BangKeoInTab(TabBase):
    def __init__(self, notebook, parent_form):
        super().__init__(parent_form)
        self.tab = ttk.Frame(notebook)
        notebook.add(self.tab, text="Băng Keo In")
        self.db_session = parent_form.db_session

        # Configure tab to expand
        self.tab.grid_rowconfigure(0, weight=1)
        self.tab.grid_columnconfigure(0, weight=1)

        # Create main frame with padding
        main_frame = ttk.Frame(self.tab, padding="10")
        main_frame.grid(row=0, column=0, sticky='nsew')

        # Configure main frame grid weights
        main_frame.grid_rowconfigure(1, weight=2)  # Basic info frame
        main_frame.grid_rowconfigure(2, weight=3)  # Price frame
        main_frame.grid_rowconfigure(3, weight=0)  # Button frame
        main_frame.grid_columnconfigure(0, weight=1)

        # Build the UI components
        self.build_ui(main_frame)
        self.bind_events()
        self.bind_currency_format()
        self.bind_shortcuts()

        # Checkbox for status
        self.da_giao = tk.BooleanVar(value=False)
        self.da_tat_toan_var = tk.BooleanVar(value=False)

    def build_ui(self, main_frame):
        # Title with better styling
        title_label = ttk.Label(main_frame, text="Băng Keo In", style='Header.TLabel')
        title_label.grid(row=0, column=0, pady=(0, 10), sticky='ew')

        # Basic Information Frame
        basic_info_frame = ttk.LabelFrame(main_frame, text="Thông tin cơ bản", padding=10)
        basic_info_frame.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
        
        # Configure grid for basic info frame
        for i in range(4):
            basic_info_frame.columnconfigure(i, weight=1)
        for i in range(4):
            basic_info_frame.rowconfigure(i, weight=1)

        # Row 0: Tên hàng và Ngày dự kiến
        ttk.Label(basic_info_frame, text="Tên hàng:").grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.ten_hang_entry = ttk.Entry(basic_info_frame)
        self.ten_hang_entry.grid(row=0, column=1, sticky='ew', padx=5, pady=5)

        ttk.Label(basic_info_frame, text="Ngày dự kiến:").grid(row=0, column=2, sticky='e', padx=5, pady=5)
        self.ngay_du_kien = DateEntry(basic_info_frame, width=20, background='darkblue', foreground='white', borderwidth=2)
        self.ngay_du_kien.grid(row=0, column=3, sticky='ew', padx=5, pady=5)

        # Row 1: Tên khách hàng
        ttk.Label(basic_info_frame, text="Tên khách hàng:").grid(row=1, column=0, sticky='e', padx=5, pady=5)
        self.ten_khach_hang_entry = ttk.Entry(basic_info_frame)
        self.ten_khach_hang_entry.grid(row=1, column=1, sticky='ew', padx=5, pady=5)

        # Row 2: Quy cách
        ttk.Label(basic_info_frame, text="Quy cách (mm):").grid(row=2, column=0, sticky='e', padx=5, pady=5)
        self.quy_cach_mm = ttk.Entry(basic_info_frame)
        self.quy_cach_mm.grid(row=2, column=1, sticky='ew', padx=5, pady=5)

        ttk.Label(basic_info_frame, text="Quy cách (m):").grid(row=2, column=2, sticky='e', padx=5, pady=5)
        self.quy_cach_m = ttk.Entry(basic_info_frame)
        self.quy_cach_m.grid(row=2, column=3, sticky='ew', padx=5, pady=5)

        # Row 3: Quy cách mic và cuộn/cây
        ttk.Label(basic_info_frame, text="Quy cách (mic):").grid(row=3, column=0, sticky='e', padx=5, pady=5)
        self.quy_cach_mic = ttk.Entry(basic_info_frame)
        self.quy_cach_mic.grid(row=3, column=1, sticky='ew', padx=5, pady=5)

        ttk.Label(basic_info_frame, text="Cuộn/Cây:").grid(row=3, column=2, sticky='e', padx=5, pady=5)
        self.cuon_cay = ttk.Entry(basic_info_frame)
        self.cuon_cay.grid(row=3, column=3, sticky='ew', padx=5, pady=5)

        # Price Frame
        price_frame = ttk.LabelFrame(main_frame, text="Giá và Chi phí", padding=10)
        price_frame.grid(row=2, column=0, sticky='nsew', padx=5, pady=5)
        
        # Configure grid for price frame
        for i in range(4):
            price_frame.columnconfigure(i, weight=1)
        for i in range(11):  # Increased to accommodate all rows
            price_frame.rowconfigure(i, weight=1)

        # Row 0: Số Lượng và Phí số lượng
        ttk.Label(price_frame, text="Số Lượng:").grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.so_luong = ttk.Entry(price_frame, width=20)
        self.so_luong.grid(row=0, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(price_frame, text="Phí số lượng:").grid(row=0, column=2, sticky='e', padx=5, pady=5)
        self.phi_sl = ttk.Entry(price_frame, width=20)
        self.phi_sl.grid(row=0, column=3, sticky='w', padx=5, pady=5)

        # Row 1: Màu Keo và Phí Keo
        ttk.Label(price_frame, text="Màu Keo:").grid(row=1, column=0, sticky='e', padx=5, pady=5)
        self.mau_keo = ttk.Entry(price_frame, width=20)
        self.mau_keo.grid(row=1, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(price_frame, text="Phí Keo:").grid(row=1, column=2, sticky='e', padx=5, pady=5)
        self.phi_keo = ttk.Entry(price_frame, width=20)
        self.phi_keo.grid(row=1, column=3, sticky='w', padx=5, pady=5)

        # Row 2: Màu Sắc và Phí màu
        ttk.Label(price_frame, text="Màu Sắc:").grid(row=2, column=0, sticky='e', padx=5, pady=5)
        self.mau_sac = ttk.Entry(price_frame, width=20)
        self.mau_sac.grid(row=2, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(price_frame, text="Phí màu:").grid(row=2, column=2, sticky='e', padx=5, pady=5)
        self.phi_mau = ttk.Entry(price_frame, width=20)
        self.phi_mau.grid(row=2, column=3, sticky='w', padx=5, pady=5)

        # Row 3: Phí size và Phí cắt
        ttk.Label(price_frame, text="Phí size:").grid(row=3, column=0, sticky='e', padx=5, pady=5)
        self.phi_size = ttk.Entry(price_frame, width=20)
        self.phi_size.grid(row=3, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(price_frame, text="Phí cắt:").grid(row=3, column=2, sticky='e', padx=5, pady=5)
        self.phi_cat = ttk.Entry(price_frame, width=20)
        self.phi_cat.grid(row=3, column=3, sticky='w', padx=5, pady=5)

        # Row 4: Đơn giá vốn và Đơn giá bán
        ttk.Label(price_frame, text="Đơn giá vốn:").grid(row=4, column=0, sticky='e', padx=5, pady=5)
        self.don_gia_von = ttk.Entry(price_frame, width=20)
        self.don_gia_von.grid(row=4, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(price_frame, text="Đơn giá(bán):").grid(row=4, column=2, sticky='e', padx=5, pady=5)
        self.don_gia_ban = ttk.Entry(price_frame, width=20)
        self.don_gia_ban.grid(row=4, column=3, sticky='w', padx=5, pady=5)

        # Row 5: Đơn giá gốc và Thành tiền bán
        ttk.Label(price_frame, text="Đơn giá gốc:").grid(row=5, column=0, sticky='e', padx=5, pady=5)
        self.don_gia_goc = ttk.Entry(price_frame, width=20, state='readonly')
        self.don_gia_goc.grid(row=5, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(price_frame, text="Thành tiền(bán):").grid(row=5, column=2, sticky='e', padx=5, pady=5)
        self.thanh_tien_ban = ttk.Entry(price_frame, width=20, state='readonly')
        self.thanh_tien_ban.grid(row=5, column=3, sticky='w', padx=5, pady=5)

        # Row 6: Thành tiền gốc và Tiền cọc
        ttk.Label(price_frame, text="Thành tiền(gốc):").grid(row=6, column=0, sticky='e', padx=5, pady=5)
        self.thanh_tien_goc = ttk.Entry(price_frame, width=20, state='readonly')
        self.thanh_tien_goc.grid(row=6, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(price_frame, text="Tiền cọc:").grid(row=6, column=2, sticky='e', padx=5, pady=5)
        self.tien_coc = ttk.Entry(price_frame, width=20)
        self.tien_coc.grid(row=6, column=3, sticky='w', padx=5, pady=5)

        # Row 7: Công nợ khách và Tiền ship
        ttk.Label(price_frame, text="Công nợ khách:").grid(row=7, column=0, sticky='e', padx=5, pady=5)
        self.cong_no_khach = ttk.Entry(price_frame, width=20, state='readonly')
        self.cong_no_khach.grid(row=7, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(price_frame, text="Tiền ship:").grid(row=7, column=2, sticky='e', padx=5, pady=5)
        self.tien_ship = ttk.Entry(price_frame, width=20)
        self.tien_ship.grid(row=7, column=3, sticky='w', padx=5, pady=5)

        # Row 8: CTV và Hoa hồng
        ttk.Label(price_frame, text="CTV:").grid(row=8, column=0, sticky='e', padx=5, pady=5)
        self.ctv = ttk.Entry(price_frame, width=20)
        self.ctv.grid(row=8, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(price_frame, text="Hoa hồng(%):").grid(row=8, column=2, sticky='e', padx=5, pady=5)
        self.hoa_hong = ttk.Entry(price_frame, width=20)
        self.hoa_hong.grid(row=8, column=3, sticky='w', padx=5, pady=5)

        # Row 9: Tiền hoa hồng và Lợi nhuận
        ttk.Label(price_frame, text="Tiền hoa hồng:").grid(row=9, column=0, sticky='e', padx=5, pady=5)
        self.tien_hoa_hong = ttk.Entry(price_frame, width=20, state='readonly')
        self.tien_hoa_hong.grid(row=9, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(price_frame, text="Lợi nhuận:").grid(row=9, column=2, sticky='e', padx=5, pady=5)
        self.loi_nhuan = ttk.Entry(price_frame, width=20, state='readonly')
        self.loi_nhuan.grid(row=9, column=3, sticky='w', padx=5, pady=5)

        # Row 10: Lõi Giấy và Thùng/Bao
        ttk.Label(price_frame, text="Lõi Giấy:").grid(row=10, column=0, sticky='e', padx=5, pady=5)
        self.loi_giay = ttk.Entry(price_frame, width=20)
        self.loi_giay.grid(row=10, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(price_frame, text="Thùng/Bao:").grid(row=10, column=2, sticky='e', padx=5, pady=5)
        self.thung_bao = ttk.Entry(price_frame, width=20)
        self.thung_bao.grid(row=10, column=3, sticky='w', padx=5, pady=5)

        # Row 11: Lợi nhuận ròng
        ttk.Label(price_frame, text="Lợi nhuận ròng:").grid(row=11, column=0, sticky='e', padx=5, pady=5)
        self.loi_nhuan_rong = ttk.Entry(price_frame, width=20, state='readonly')
        self.loi_nhuan_rong.grid(row=11, column=1, sticky='w', padx=5, pady=5)

        # Buttons Frame
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=3, column=0, pady=10, sticky='ew')

        # Configure button frame columns
        for i in range(5):  # 5 columns for 5 buttons
            button_frame.columnconfigure(i, weight=1)

        # Create buttons with consistent styling
        buttons = [
            ("Tính toán", self.tinh_toan),
            ("Lưu", self.luu_don_hang),
            ("Xuất Excel", self.export_to_excel),
            ("Xuất Email", self.export_email),
            ("Xóa", self.xoa_form)
        ]

        for i, (text, command) in enumerate(buttons):
            btn = ttk.Button(button_frame, text=text, command=command, style='Custom.TButton', width=15)
            btn.grid(row=0, column=i, padx=5)

        # Set focus to the first entry
        self.ten_hang_entry.focus_set()

    def _configure_grid(self, frame, cols):
        """Configure grid columns to be evenly spaced"""
        for i in range(cols):
            frame.columnconfigure(i, weight=1)

    def bind_events(self):
        entries_to_bind = [
            self.so_luong, self.phi_sl, self.phi_keo, self.phi_size,
            self.phi_cat, self.don_gia_von, self.don_gia_ban, self.tien_coc,
            self.hoa_hong, self.phi_mau, self.quy_cach_mm,
            self.quy_cach_m, self.quy_cach_mic, self.cuon_cay,
            self.tien_ship  # Thêm tiền ship
        ]
        for entry in entries_to_bind:
            entry.bind('<KeyRelease>', self.auto_calculate)

        # Register validation command
        vcmd = (self.root.register(self.is_valid_float), '%P')
        entries_to_validate = entries_to_bind
        for entry in entries_to_validate:
            entry.config(validate='key', validatecommand=vcmd)

    def auto_calculate(self, event):
        self.tinh_toan()

    def bind_currency_format(self):
        currency_fields = [
            self.don_gia_von,
            self.don_gia_ban,
            self.tien_coc,
            self.phi_sl,
            self.phi_keo,
            self.phi_mau,
            self.phi_size,
            self.phi_cat,
        ]
        for field in currency_fields:
            field.bind('<FocusOut>', self.format_currency_input)

    def bind_shortcuts(self):
        self.root.bind('<Control-s>', lambda event: self.luu_don_hang())
        self.root.bind('<Control-t>', lambda event: self.tinh_toan())
        self.root.bind('<Control-e>', lambda event: self.export_to_excel())
        self.root.bind('<Control-q>', lambda event: self.thoat())

    def tinh_toan(self):
        try:
            # 1. Các giá trị đầu vào cơ bản
            don_gia_von = self.validate_float_input(self.don_gia_von.get())
            phi_sl = self.validate_float_input(self.phi_sl.get())
            phi_mau = self.validate_float_input(self.phi_mau.get())
            phi_keo = self.validate_float_input(self.phi_keo.get())
            phi_size = self.validate_float_input(self.phi_size.get())
            phi_cat = self.validate_float_input(self.phi_cat.get())
            so_luong = self.validate_float_input(self.so_luong.get())
            don_gia_ban = self.validate_float_input(self.don_gia_ban.get())
            tien_coc = self.validate_float_input(self.tien_coc.get())
            hoa_hong = self.validate_float_input(self.hoa_hong.get()) / 100  # Chuyển % thành decimal
            tien_ship = self.validate_float_input(self.tien_ship.get())

            # 2. Tính đơn giá gốc
            cuon_cay = self.validate_float_input(self.cuon_cay.get())
            quy_cach_m = self.validate_float_input(self.quy_cach_m.get())

            if cuon_cay == 0 or quy_cach_m == 0:
                don_gia_goc = 0
            else:
                # Công thức tính đơn giá gốc:
                # (Tổng chi phí) / 90 * quy_cach_m / cuon_cay
                don_gia_goc = (don_gia_von + phi_sl + phi_mau + phi_keo + phi_size + phi_cat) / 90 * quy_cach_m / cuon_cay

            # 3. Tính các giá trị phái sinh
            thanh_tien_goc = don_gia_goc * so_luong  # Thành tiền gốc = đơn giá gốc × số lượng
            thanh_tien_ban = don_gia_ban * so_luong  # Thành tiền bán = đơn giá bán × số lượng
            cong_no_khach = thanh_tien_ban - tien_coc  # Công nợ = thành tiền bán - tiền cọc
            loi_nhuan = thanh_tien_ban - thanh_tien_goc  # Lợi nhuận = thành tiền bán - thành tiền gốc
            tien_hoa_hong = loi_nhuan * hoa_hong  # Tiền hoa hồng = lợi nhuận × tỷ lệ hoa hồng
            loi_nhuan_rong = loi_nhuan - tien_hoa_hong - tien_ship  # Lợi nhuận ròng = lợi nhuận - tiền hoa hồng - tiền ship

            # Cập nhật giao diện
            self.update_readonly_field(self.don_gia_goc, don_gia_goc)
            self.update_readonly_field(self.thanh_tien_goc, thanh_tien_goc)
            self.update_readonly_field(self.thanh_tien_ban, thanh_tien_ban)
            self.update_readonly_field(self.cong_no_khach, cong_no_khach)
            self.update_readonly_field(self.tien_hoa_hong, tien_hoa_hong)
            self.update_readonly_field(self.loi_nhuan, loi_nhuan)
            self.update_readonly_field(self.loi_nhuan_rong, loi_nhuan_rong)

            # Update status bar
            self.update_status("Tính toán thành công")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi tính toán: {str(e)}")
            self.update_status("Lỗi khi tính toán")

    def luu_don_hang(self):
        """Save the order to database"""
        try:
            # Kiểm tra các trường bắt buộc
            required_fields = {
                'Tên hàng': self.ten_hang_entry.get(),
                'Tên khách hàng': self.ten_khach_hang_entry.get(),
                'Số lượng': self.so_luong.get(),
                'Đơn giá bán': self.don_gia_ban.get()
            }
            
            # Validate required fields
            for field_name, value in required_fields.items():
                if not value:
                    messagebox.showerror("Lỗi", f"Vui lòng nhập {field_name}")
                    return
            
            # Create data dictionary
            data = {
                'thoi_gian': datetime.now(),
                'ten_hang': self.ten_hang_entry.get(),
                'ten_khach_hang': self.ten_khach_hang_entry.get(),
                'ngay_du_kien': self.ngay_du_kien.get_date(),
                'quy_cach_mm': self.parse_float(self.quy_cach_mm.get()),
                'quy_cach_m': self.parse_float(self.quy_cach_m.get()),
                'quy_cach_mic': self.parse_float(self.quy_cach_mic.get()),
                'cuon_cay': self.parse_float(self.cuon_cay.get()),
                'so_luong': self.parse_float(self.so_luong.get()),
                'phi_sl': self.parse_float(self.phi_sl.get()),
                'mau_keo': self.mau_keo.get(),
                'phi_keo': self.parse_float(self.phi_keo.get()),
                'mau_sac': self.mau_sac.get(),
                'phi_mau': self.parse_float(self.phi_mau.get()),
                'phi_size': self.parse_float(self.phi_size.get()),
                'phi_cat': self.parse_float(self.phi_cat.get()),
                'don_gia_von': self.parse_float(self.don_gia_von.get()),
                'don_gia_goc': self.parse_float(self.don_gia_goc.get()),
                'thanh_tien_goc': self.parse_float(self.thanh_tien_goc.get()),
                'don_gia_ban': self.parse_float(self.don_gia_ban.get()),
                'thanh_tien_ban': self.parse_float(self.thanh_tien_ban.get()),
                'tien_coc': self.parse_float(self.tien_coc.get()),
                'cong_no_khach': self.parse_float(self.cong_no_khach.get()),
                'ctv': self.ctv.get(),
                'hoa_hong': self.parse_float(self.hoa_hong.get()),
                'tien_hoa_hong': self.parse_float(self.tien_hoa_hong.get()),
                'loi_giay': self.loi_giay.get(),
                'thung_bao': self.thung_bao.get(),
                'loi_nhuan': self.parse_float(self.loi_nhuan.get()),
                'tien_ship': self.parse_float(self.tien_ship.get()),
                'loi_nhuan_rong': self.parse_float(self.loi_nhuan_rong.get()),
                'da_giao': False,
                'da_tat_toan': False
            }
            
            # Tạo đối tượng BangKeoInOrder mới
            don_hang = BangKeoInOrder(**data)
            
            # Thêm vào database
            self.db_session.add(don_hang)
            self.db_session.commit()
            
            # Hiển thị thông báo thành công với mã đơn hàng
            messagebox.showinfo("Thành công", f"Đã lưu đơn hàng thành công!\nMã đơn hàng: {don_hang.id}")
            
            # Cập nhật history tab và thống kê tab
            if hasattr(self.parent_form, 'history_tab'):
                self.parent_form.history_tab.refresh_data()
            if hasattr(self.parent_form, 'thong_ke_tab'):
                self.parent_form.thong_ke_tab.load_data()
                
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lưu đơn hàng: {str(e)}")
            self.db_session.rollback()

    def xoa_form(self):
        """Clear all form fields"""
        try:
            # Clear basic information
            self.ten_hang_entry.delete(0, tk.END)
            self.ten_khach_hang_entry.delete(0, tk.END)
            self.quy_cach_mm.delete(0, tk.END)
            self.quy_cach_m.delete(0, tk.END)
            self.quy_cach_mic.delete(0, tk.END)
            self.cuon_cay.delete(0, tk.END)
            
            # Clear quantities and fees
            self.so_luong.delete(0, tk.END)
            self.phi_sl.delete(0, tk.END)
            self.mau_keo.delete(0, tk.END)
            self.phi_keo.delete(0, tk.END)
            self.mau_sac.delete(0, tk.END)
            self.phi_mau.delete(0, tk.END)
            self.phi_size.delete(0, tk.END)
            self.phi_cat.delete(0, tk.END)
            
            # Clear prices
            self.don_gia_von.delete(0, tk.END)
            self.don_gia_goc.delete(0, tk.END)
            self.thanh_tien_goc.delete(0, tk.END)
            self.don_gia_ban.delete(0, tk.END)
            self.thanh_tien_ban.delete(0, tk.END)
            self.tien_coc.delete(0, tk.END)
            self.cong_no_khach.delete(0, tk.END)
            
            # Clear CTV and commission
            self.ctv.delete(0, tk.END)
            self.hoa_hong.delete(0, tk.END)
            self.tien_hoa_hong.delete(0, tk.END)
            
            # Clear additional info
            self.loi_giay.delete(0, tk.END)
            self.thung_bao.delete(0, tk.END)
            self.loi_nhuan.delete(0, tk.END)
            self.tien_ship.delete(0, tk.END)
            self.loi_nhuan_rong.delete(0, tk.END)
            
            # Reset date to today
            self.ngay_du_kien.set_date(datetime.now())
            
            # Set focus to first field
            self.ten_hang_entry.focus_set()
            
            messagebox.showinfo("Thành công", "Đã xóa form")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi xóa form: {str(e)}")
            raise

    def thoat(self):
        try:
            if messagebox.askyesno("Xác nhận", "Bạn có chắc muốn thoát chương trình?"):
                self.root.quit()
                self.update_status("Thoát chương trình")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi thoát: {str(e)}")
            self.update_status("Lỗi khi thoát chương trình")

    def export_to_excel(self):
        """Export the order details to Excel"""
        try:
            current_date = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_name = f"don_hang_{current_date}.xlsx"
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Đơn hàng"
            
            # Prepare data
            data = {
                'ID': self.ten_hang_entry.get(),
                'Ngày': current_date,
                'Tên Hàng': self.ten_hang_entry.get(),
                'Tên Khách Hàng': self.ten_khach_hang_entry.get(),
                'Ngày dự kiến': self.ngay_du_kien.get_date().strftime('%d-%m-%Y'),
                'Quy Cách (mm)': self.quy_cach_mm.get(),
                'Quy Cách (m)': self.quy_cach_m.get(),
                'Quy Cách (mic)': self.quy_cach_mic.get(),
                'Cuộn/Cây': self.cuon_cay.get(),
                'Số lượng': self.so_luong.get(),
                'Phí SL': self.phi_sl.get(),
                'Màu keo': self.mau_keo.get(),
                'Phí keo': self.phi_keo.get(),
                'Màu sắc': self.mau_sac.get(),
                'Phí màu': self.phi_mau.get(),
                'Phí size': self.phi_size.get(),
                'Phí cắt': self.phi_cat.get(),
                'Đơn giá vốn': self.don_gia_von.get(),
                'Đơn giá gốc': self.don_gia_goc.get(),
                'Thành tiền gốc': self.thanh_tien_goc.get(),
                'Đơn giá bán': self.don_gia_ban.get(),
                'Thành tiền bán': self.thanh_tien_ban.get(),
                'Tiền cọc': self.tien_coc.get(),
                'Công nợ khách': self.cong_no_khach.get(),
                'CTV': self.ctv.get(),
                'Hoa hồng': self.hoa_hong.get(),
                'Tiền hoa hồng': self.tien_hoa_hong.get(),
                'Lỗi giấy': self.loi_giay.get(),
                'Thùng/Bao': self.thung_bao.get(),
                'Lợi nhuận': self.loi_nhuan.get(),
                'Tiền ship': self.tien_ship.get(),
                'Lợi nhuận ròng': self.loi_nhuan_rong.get()
            }
            
            # Write headers and data
            for col, (header, value) in enumerate(data.items(), start=1):
                ws.cell(row=1, column=col, value=header)
                ws.cell(row=2, column=col, value=value)
            
            # Save file
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=file_name
            )
            
            if file_path:
                wb.save(file_path)
                messagebox.showinfo("Thành công", f"Đã xuất file Excel: {file_path}")
        
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi xuất Excel: {str(e)}")
            raise

    def export_email(self):
        """Export the order details for email"""
        try:
            ten_hang = self.ten_hang_entry.get()
            ten_khach_hang = self.ten_khach_hang_entry.get()
            mau_sac = self.mau_sac.get()
            mau_keo = self.mau_keo.get()
            quy_cach = f"{self.quy_cach_mm.get()}mm x {self.quy_cach_m.get()}m x {self.quy_cach_mic.get()}mic"
            so_luong = self.so_luong.get()
            loi_giay = self.loi_giay.get()
            thung_bao = self.thung_bao.get()
            
            content = f"""
Chào bác,

Bác làm giúp con đơn hàng in logo bên dưới nhé:

THÔNG TIN ĐƠN HÀNG BĂNG KEO IN:
---------------------------
Tên hàng: {ten_hang}
Tên khách hàng: {ten_khach_hang}
Màu sắc: {mau_sac}
Màu keo: {mau_keo}
Quy cách: {quy_cach}
Số lượng: {so_luong}
Lõi giấy: {loi_giay}
Thùng/Bao: {thung_bao}

Cảm ơn bác!
Quế
"""
            
            # Create temporary file
            current_date = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_name = f"bang_keo_in_{current_date}.txt"
            temp_file = os.path.join(os.environ.get('TEMP') or os.environ.get('TMP') or '/tmp', file_name)
            
            # Write content to temp file
            with open(temp_file, 'w', encoding='utf-8') as f:
                f.write(content)
            
            # Open the file with default text editor
            os.startfile(temp_file)
            
            # Also allow saving to custom location
            file_path = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt")],
                initialfile=file_name
            )
            
            if file_path:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(content)
                messagebox.showinfo("Thành công", f"Đã xuất file: {file_path}")
        
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi xuất file: {str(e)}")
            raise
