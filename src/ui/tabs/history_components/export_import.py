import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
import logging
from src.ui.tabs.history_components.email_dialog import EmailDialog

class ExportImportManager:
    def __init__(self, parent):
        self.parent = parent
        self.DATE_FORMAT = '%d/%m/%Y'
    
    @staticmethod
    def format_number(value):
        """Format số: nếu là số nguyên thì không hiển thị .0, chỉ hiển thị phần thập phân nếu có"""
        if value is None or value == '':
            return ''
        try:
            num = float(value)
            if num.is_integer():
                return str(int(num))
            return str(num)
        except (ValueError, TypeError):
            return str(value)
        
    def create_import_export_buttons(self, button_frame):
        button_frame.columnconfigure((0, 1, 2, 3), weight=1)
    
        export_bang_keo_in_button = ttk.Button(button_frame, text="Export Băng Keo Template", 
                                              command=lambda: self.export_template('bang_keo_in'))
        export_bang_keo_in_button.grid(row=0, column=0, padx=5, sticky='ew')
    
        export_truc_in_button = ttk.Button(button_frame, text="Export Trục In Template", 
                                         command=lambda: self.export_template('truc_in'))
        export_truc_in_button.grid(row=0, column=1, padx=5, sticky='ew')
    
        import_bang_keo_in_button = ttk.Button(button_frame, text="Import Băng Keo Data", 
                                             command=lambda: self.import_data('bang_keo_in'))
        import_bang_keo_in_button.grid(row=0, column=2, padx=5, sticky='ew')
    
        import_truc_in_button = ttk.Button(button_frame, text="Import Trục In Data", 
                                         command=lambda: self.import_data('truc_in'))
        import_truc_in_button.grid(row=0, column=3, padx=5, sticky='ew')
        
    def export_selected_to_excel(self, tree, sheet_name):
        try:
            selected_items = tree.selection()
            if not selected_items:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn ít nhất một dòng để xuất Excel")
                return
            
            headers = [tree.heading(col)['text'] for col in tree['columns']]
            
            data = []
            for item in selected_items:
                values = list(tree.item(item)['values'])
                if sheet_name == "Bang keo in":
                    if values[1]:  # thoi_gian
                        try:
                            date_obj = datetime.strptime(values[1], self.DATE_FORMAT)
                            values[1] = date_obj.strftime('%m/%d/%Y')
                        except ValueError:
                            pass
                    if values[3]:  # ngay_du_kien
                        try:
                            date_obj = datetime.strptime(values[3], self.DATE_FORMAT)
                            values[3] = date_obj.strftime('%m/%d/%Y')
                        except ValueError:
                            pass
                else:
                    if values[1]:  # thoi_gian
                        try:
                            date_obj = datetime.strptime(values[1], self.DATE_FORMAT)
                            values[1] = date_obj.strftime('%m/%d/%Y')
                        except ValueError:
                            pass
                    if values[3]:  # ngay_du_kien
                        try:
                            date_obj = datetime.strptime(values[3], self.DATE_FORMAT)
                            values[3] = date_obj.strftime('%m/%d/%Y')
                        except ValueError:
                            pass
                data.append(values)
            
            current_time = datetime.now().strftime('%d-%m-%y')
            default_filename = f"DonHang_{sheet_name}_{current_time}.xlsx"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=default_filename
            )
            
            if not file_path:
                return

            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    next_row = ws.max_row + 1
                else:
                    ws = wb.create_sheet(sheet_name)
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)
                    next_row = 2
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                next_row = 2

            for row_data in data:
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=next_row, column=col, value=value)
                next_row += 1

            wb.save(file_path)
            
            messagebox.showinfo("Thành công", f"Đã xuất dữ liệu ra file Excel:\n{file_path}")
            self.parent.update_status("Đã xuất Excel thành công")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi xuất Excel: {str(e)}")
            self.parent.update_status("Lỗi khi xuất Excel")
            
    def export_selected_to_email(self, tree, order_type):
        try:
            selected_items = tree.selection()
            if not selected_items:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn ít nhất một dòng để gửi email")
                return

            values = tree.item(selected_items[0])['values']
            order_id = values[0]

            # Fetch order from database to ensure we have all fields
            order = None
            if order_type == 'bang_keo_in':
                from src.database.database import BangKeoInOrder
                order = self.parent.db_session.query(BangKeoInOrder).filter_by(id=order_id).first()
            elif order_type == 'truc_in':
                from src.database.database import TrucInOrder
                order = self.parent.db_session.query(TrucInOrder).filter_by(id=order_id).first()
            elif order_type == 'bang_keo':
                from src.database.database import BangKeoOrder
                order = self.parent.db_session.query(BangKeoOrder).filter_by(id=order_id).first()
                
            if not order:
                messagebox.showerror("Lỗi", "Không tìm thấy thông tin đơn hàng trong cơ sở dữ liệu.")
                return

            # Build subject and body templates
            if order_type == 'bang_keo_in':
                quy_cach_mm = self.format_number(order.quy_cach_mm) if order.quy_cach_mm else "0"
                quy_cach_m = self.format_number(order.quy_cach_m) if order.quy_cach_m else "0"
                quy_cach_mic = self.format_number(order.quy_cach_mic) if order.quy_cach_mic else "0"
                so_luong = self.format_number(order.so_luong) if order.so_luong else "0"
                subject = f"Băng keo in {order.ten_hang}"
                email_content = (
                    f"Chào bác,\n\n"
                    f"Bác làm giúp con đơn hàng in logo \"{order.ten_hang}\" này nhé\n"
                    f"Thông tin đơn hàng:\n"
                    f"________________________________\n"
                    f"Màu sắc: {order.mau_sac}\n"
                    f"Màu keo: {order.mau_keo}\n"
                    f"Số lượng: {so_luong} cuộn\n"
                    f"Quy cách: {quy_cach_mm}mm * {quy_cach_m}m * {quy_cach_mic}mic\n"
                    f"Lõi giấy: {order.loi_giay}\n"
                    f"Thùng bao: {order.thung_bao}\n"
                    f"________________________________\n\n"
                    f"Cám ơn bác\n"
                    f"Quế"
                )
            elif order_type == 'truc_in':
                so_luong = self.format_number(order.so_luong) if order.so_luong else "0"
                subject = f"Trục in {order.ten_hang}"
                email_content = (
                    f"Chào bác,\n\n"
                    f"Bác làm giúp con đơn hàng Trục In \"{order.ten_hang}\" này nhé\n"
                    f"Thông tin đơn hàng:\n"
                    f"________________________________\n"
                    f"Màu sắc: {order.mau_sac}\n"
                    f"Màu keo: {order.mau_keo}\n"
                    f"Số lượng: {so_luong} cuộn\n"
                    f"Quy cách: {order.quy_cach}\n"
                    f"________________________________\n\n"
                    f"Cám ơn bác\n"
                    f"Quế"
                )
            else:
                so_luong = self.format_number(order.so_luong) if order.so_luong else "0"
                quy_cach = self.format_number(order.quy_cach) if order.quy_cach else ""
                subject = f"Băng keo {order.ten_hang}"
                email_content = (
                    f"Chào bác,\n\n"
                    f"Bác làm giúp con đơn hàng Băng Keo \"{order.ten_hang}\" này nhé\n"
                    f"Thông tin đơn hàng:\n"
                    f"________________________________\n"
                    f"Màu sắc: {order.mau_sac}\n"
                    f"Số lượng: {so_luong} KG\n"
                    f"Quy cách: {quy_cach} KG\n\n"
                    f"________________________________\n\n"
                    f"Cám ơn bác\n"
                    f"Quế"
                )

            dlg = EmailDialog(self.parent.root, self.parent.db_session, order_type, order_id, subject, email_content)
            self.parent.root.wait_window(dlg)

        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi chuẩn bị email: {str(e)}")
            self.parent.update_status("Lỗi khi gửi email") 