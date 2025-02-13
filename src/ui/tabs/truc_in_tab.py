# truc_in_tab.py
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from src.ui.tabs.tab_base import TabBase
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from tkcalendar import DateEntry
from src.database.database import TrucInOrder

class TrucInTab(TabBase):
    def __init__(self, notebook, parent_form):
        super().__init__(parent_form)
        self.tab = ttk.Frame(notebook)
        notebook.add(self.tab, text="Trục In")
        self.db_session = parent_form.db_session

        # Create main frame with padding
        main_frame = ttk.Frame(self.tab, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Configure the grid to expand properly
        main_frame.columnconfigure(0, weight=1)
        for row_index in range(4):  # Assuming 4 rows including buttons
            if row_index == 3:
                main_frame.rowconfigure(row_index, weight=0)  # Buttons row should not expand
            else:
                main_frame.rowconfigure(row_index, weight=1)

        # Build the UI components
        self.build_ui(main_frame)
        self.bind_events()
        self.bind_currency_format()
        self.bind_shortcuts()

        # Checkbox for status
        self.da_giao = tk.BooleanVar(value=False)
        self.da_tat_toan_var = tk.BooleanVar(value=False)

    def build_ui(self, main_frame):
        # Configure grid columns
        main_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.columnconfigure(2, weight=1)
        main_frame.columnconfigure(3, weight=1)

        # Title
        title_label = ttk.Label(main_frame, text="Trục In", font=('Segoe UI', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=4, pady=(0, 20), sticky='ew')

        # Basic Information Frame
        basic_info_frame = ttk.LabelFrame(main_frame, text="Thông tin cơ bản", padding=10)
        basic_info_frame.grid(row=1, column=0, columnspan=4, sticky='nsew', padx=5, pady=5)
        self._configure_grid(basic_info_frame, 4)

        # Row 0: Tên hàng và Ngày dự kiến
        ttk.Label(basic_info_frame, text="Tên hàng:").grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.ten_hang_entry = ttk.Entry(basic_info_frame)
        self.ten_hang_entry.grid(row=0, column=1, sticky='ew', padx=5, pady=5)

        ttk.Label(basic_info_frame, text="Ngày dự kiến:").grid(row=0, column=2, sticky='e', padx=5, pady=5)
        self.ngay_du_kien = DateEntry(basic_info_frame, width=20, background='darkblue', foreground='white', borderwidth=2)
        self.ngay_du_kien.grid(row=0, column=3, sticky='ew', padx=5, pady=5)

        # Row 1: Tên khách hàng
        ttk.Label(basic_info_frame, text="Tên khách hàng:").grid(row=1, column=0, sticky='e', padx=5, pady=5)
        self.ten_khach_hang_entry = ttk.Entry(basic_info_frame)
        self.ten_khach_hang_entry.grid(row=1, column=1, sticky='ew', padx=5, pady=5)

        # Row 2: Quy cách và Số lượng
        ttk.Label(basic_info_frame, text="Quy cách:").grid(row=2, column=0, sticky='e', padx=5, pady=5)
        self.quy_cach = ttk.Entry(basic_info_frame)
        self.quy_cach.grid(row=2, column=1, sticky='ew', padx=5, pady=5)

        ttk.Label(basic_info_frame, text="Số lượng:").grid(row=2, column=2, sticky='e', padx=5, pady=5)
        self.so_luong = ttk.Entry(basic_info_frame)
        self.so_luong.grid(row=2, column=3, sticky='ew', padx=5, pady=5)

        # Row 3: Màu sắc và Màu keo
        ttk.Label(basic_info_frame, text="Màu sắc:").grid(row=3, column=0, sticky='e', padx=5, pady=5)
        self.mau_sac = ttk.Entry(basic_info_frame)
        self.mau_sac.grid(row=3, column=1, sticky='ew', padx=5, pady=5)

        ttk.Label(basic_info_frame, text="Màu keo:").grid(row=3, column=2, sticky='e', padx=5, pady=5)
        self.mau_keo = ttk.Entry(basic_info_frame)
        self.mau_keo.grid(row=3, column=3, sticky='ew', padx=5, pady=5)

        # Giá cả Frame
        price_frame = ttk.LabelFrame(main_frame, text="Giá cả", padding=10)
        price_frame.grid(row=2, column=0, columnspan=4, sticky='nsew', padx=5, pady=5)
        self._configure_grid(price_frame, 4)

        # Đơn giá gốc
        ttk.Label(price_frame, text="Đơn giá gốc:").grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.truc_in_don_gia_goc = ttk.Entry(price_frame, width=20)
        self.truc_in_don_gia_goc.grid(row=0, column=1, sticky='w', padx=5, pady=5)

        # Thành tiền
        ttk.Label(price_frame, text="Thành tiền:").grid(row=1, column=0, sticky='e', padx=5, pady=5)
        self.truc_in_thanh_tien = ttk.Entry(price_frame, width=20, state='readonly')
        self.truc_in_thanh_tien.grid(row=1, column=1, sticky='w', padx=5, pady=5)

        # Đơn giá bán
        ttk.Label(price_frame, text="Đơn giá (bán):").grid(row=0, column=2, sticky='e', padx=5, pady=5)
        self.truc_in_don_gia_ban = ttk.Entry(price_frame, width=20)
        self.truc_in_don_gia_ban.grid(row=0, column=3, sticky='w', padx=5, pady=5)

        # Thành tiền bán
        ttk.Label(price_frame, text="Thành tiền (bán):").grid(row=1, column=2, sticky='e', padx=5, pady=5)
        self.truc_in_thanh_tien_ban = ttk.Entry(price_frame, width=20, state='readonly')
        self.truc_in_thanh_tien_ban.grid(row=1, column=3, sticky='w', padx=5, pady=5)

        # Công nợ khách
        ttk.Label(price_frame, text="Công nợ khách:").grid(row=2, column=0, sticky='e', padx=5, pady=5)
        self.truc_in_cong_no_khach = ttk.Entry(price_frame, width=20, state='readonly')
        self.truc_in_cong_no_khach.grid(row=2, column=1, sticky='w', padx=5, pady=5)

        # CTV và Hoa hồng
        ttk.Label(price_frame, text="CTV:").grid(row=3, column=0, sticky='e', padx=5, pady=5)
        self.truc_in_ctv = ttk.Entry(price_frame, width=20)
        self.truc_in_ctv.grid(row=3, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(price_frame, text="Hoa Hồng (%):").grid(row=3, column=2, sticky='e', padx=5, pady=5)
        self.truc_in_hoa_hong = ttk.Entry(price_frame, width=20)
        self.truc_in_hoa_hong.grid(row=3, column=3, sticky='w', padx=5, pady=5)

        # Tiền hoa hồng
        ttk.Label(price_frame, text="Tiền hoa hồng:").grid(row=4, column=0, sticky='e', padx=5, pady=5)
        self.truc_in_tien_hoa_hong = ttk.Entry(price_frame, width=20, state='readonly')
        self.truc_in_tien_hoa_hong.grid(row=4, column=1, sticky='w', padx=5, pady=5)

        # Lợi nhuận
        ttk.Label(price_frame, text="Lợi nhuận:").grid(row=5, column=0, sticky='e', padx=5, pady=5)
        self.truc_in_loi_nhuan = ttk.Entry(price_frame, width=20, state='readonly')
        self.truc_in_loi_nhuan.grid(row=5, column=1, sticky='w', padx=5, pady=5)

        # Tiền ship
        ttk.Label(price_frame, text="Tiền ship:").grid(row=5, column=2, sticky='e', padx=5, pady=5)
        self.truc_in_tien_ship = ttk.Entry(price_frame, width=20)
        self.truc_in_tien_ship.grid(row=5, column=3, sticky='w', padx=5, pady=5)

        # Lợi nhuận ròng
        ttk.Label(price_frame, text="Lợi nhuận ròng:").grid(row=6, column=0, sticky='e', padx=5, pady=5)
        self.truc_in_loi_nhuan_rong = ttk.Entry(price_frame, width=20, state='readonly')
        self.truc_in_loi_nhuan_rong.grid(row=6, column=1, sticky='w', padx=5, pady=5)

        # Buttons Frame
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=3, column=0, columnspan=4, pady=20, sticky='e')  # Align to the right

        # Define a consistent style for all buttons
        style = ttk.Style()
        style.configure('CustomButton.TButton',
                        font=('Segoe UI', 10),
                        padding=6)

        # Create buttons with the custom style
        btn_tinh_toan = ttk.Button(button_frame, text="Tính toán", command=self.tinh_toan_truc_in, style='CustomButton.TButton', width=12)
        btn_luu = ttk.Button(button_frame, text="Lưu", command=self.luu_truc_in, style='CustomButton.TButton', width=12)
        btn_xuat_excel = ttk.Button(button_frame, text="Xuất Excel", command=self.export_to_excel, style='CustomButton.TButton', width=12)
        btn_xuat_email = ttk.Button(button_frame, text="Xuất Email", command=self.export_truc_in_email, style='CustomButton.TButton', width=12)
        btn_xoa = ttk.Button(button_frame, text="Xóa", command=self.xoa_form_truc_in, style='CustomButton.TButton', width=12)

        # Pack buttons to the right with consistent padding
        btn_xoa.pack(side='right', padx=5)
        btn_xuat_email.pack(side='right', padx=5)
        btn_xuat_excel.pack(side='right', padx=5)
        btn_luu.pack(side='right', padx=5)
        btn_tinh_toan.pack(side='right', padx=5)

        # Set focus to the first entry
        self.ten_hang_entry.focus_set()

    def _configure_grid(self, frame, cols):
        """Configure grid columns to be evenly spaced"""
        for i in range(cols):
            frame.columnconfigure(i, weight=1)

    def bind_events(self):
        entries_to_bind = [
            self.so_luong,
            self.truc_in_don_gia_ban,
            self.truc_in_don_gia_goc,
            self.truc_in_hoa_hong,
            self.truc_in_tien_ship,
        ]
        for entry in entries_to_bind:
            entry.bind('<KeyRelease>', self.auto_calculate)

        # Register validation command
        vcmd = (self.root.register(self.is_valid_float), '%P')
        for entry in entries_to_bind:
            entry.config(validate='key', validatecommand=vcmd)

    def auto_calculate(self, event):
        self.tinh_toan_truc_in()

    def bind_currency_format(self):
        currency_fields = [
            self.truc_in_don_gia_goc,
            self.truc_in_don_gia_ban,
        ]
        for field in currency_fields:
            field.bind('<FocusOut>', self.format_currency_input)

    def bind_shortcuts(self):
        self.root.bind('<Control-s>', lambda event: self.luu_truc_in())
        self.root.bind('<Control-t>', lambda event: self.tinh_toan_truc_in())
        self.root.bind('<Control-e>', lambda event: self.export_to_excel())
        self.root.bind('<Control-q>', lambda event: self.root.quit())

    def tinh_toan_truc_in(self):
        try:
            # Get values
            so_luong = self.validate_float_input(self.so_luong.get())
            don_gia_ban = self.validate_float_input(self.truc_in_don_gia_ban.get())
            don_gia_goc = self.validate_float_input(self.truc_in_don_gia_goc.get())
            hoa_hong = self.validate_float_input(self.truc_in_hoa_hong.get()) / 100
            tien_ship = self.validate_float_input(self.truc_in_tien_ship.get())

            # Calculate values
            thanh_tien = don_gia_goc * so_luong  # Thành tiền gốc
            thanh_tien_ban = don_gia_ban * so_luong  # Thành tiền bán
            loi_nhuan = thanh_tien_ban - thanh_tien  # Lợi nhuận
            tien_hoa_hong = loi_nhuan * hoa_hong  # Tiền hoa hồng
            cong_no_khach = thanh_tien_ban  # Công nợ khách = thành tiền bán
            loi_nhuan_rong = loi_nhuan - tien_hoa_hong - tien_ship  # Lợi nhuận ròng = lợi nhuận - tiền hoa hồng - tiền ship

            # Update fields
            self.update_readonly_field(self.truc_in_thanh_tien, thanh_tien)
            self.update_readonly_field(self.truc_in_thanh_tien_ban, thanh_tien_ban)
            self.update_readonly_field(self.truc_in_cong_no_khach, cong_no_khach)
            self.update_readonly_field(self.truc_in_tien_hoa_hong, tien_hoa_hong)
            self.update_readonly_field(self.truc_in_loi_nhuan, loi_nhuan)
            self.update_readonly_field(self.truc_in_loi_nhuan_rong, loi_nhuan_rong)

            self.update_status("Tính toán Trục In thành công")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi tính toán Trục In: {str(e)}")
            self.update_status("Lỗi khi tính toán Trục In")

    def luu_truc_in(self):
        """Save the Truc In order to database"""
        try:
            # Kiểm tra các trường bắt buộc
            required_fields = {
                'Tên hàng': self.ten_hang_entry.get(),
                'Tên khách hàng': self.ten_khach_hang_entry.get(),
                'Số lượng': self.so_luong.get(),
                'Đơn giá bán': self.truc_in_don_gia_ban.get()
            }
            
            # Validate required fields
            for field_name, value in required_fields.items():
                if not value:
                    messagebox.showerror("Lỗi", f"Vui lòng nhập {field_name}")
                    return
            
            # Format quy_cach to include unit if not present
            quy_cach = self.quy_cach.get().strip()
            if quy_cach and not any(unit in quy_cach.lower() for unit in ['mm', 'cm', 'm']):
                quy_cach = f"{quy_cach}mm"
            
            # Create data dictionary
            data = {
                'thoi_gian': datetime.now(),
                'ten_hang': self.ten_hang_entry.get(),
                'ten_khach_hang': self.ten_khach_hang_entry.get(),
                'ngay_du_kien': self.ngay_du_kien.get_date(),
                'quy_cach': quy_cach,  # Use formatted quy_cach
                'so_luong': self.parse_float(self.so_luong.get()),
                'mau_sac': self.mau_sac.get(),
                'mau_keo': self.mau_keo.get(),
                'don_gia_goc': self.parse_float(self.truc_in_don_gia_goc.get()),
                'thanh_tien_goc': self.parse_float(self.truc_in_thanh_tien.get()),
                'don_gia_ban': self.parse_float(self.truc_in_don_gia_ban.get()),
                'thanh_tien_ban': self.parse_float(self.truc_in_thanh_tien_ban.get()),
                'cong_no_khach': self.parse_float(self.truc_in_cong_no_khach.get()),
                'ctv': self.truc_in_ctv.get(),
                'hoa_hong': self.parse_float(self.truc_in_hoa_hong.get()),
                'tien_hoa_hong': self.parse_float(self.truc_in_tien_hoa_hong.get()),
                'loi_nhuan': self.parse_float(self.truc_in_loi_nhuan.get()),
                'tien_ship': self.parse_float(self.truc_in_tien_ship.get()),
                'loi_nhuan_rong': self.parse_float(self.truc_in_loi_nhuan_rong.get()),
                'da_giao': False,
                'da_tat_toan': False
            }
            
            # Create new TrucInOrder object
            don_hang = TrucInOrder(**data)
            
            # Add to database
            self.db_session.add(don_hang)
            self.db_session.commit()
            
            messagebox.showinfo("Thành công", "Đã lưu đơn hàng thành công!")
            
            # Update history tab and statistics tab
            if hasattr(self.parent_form, 'history_tab'):
                self.parent_form.history_tab.refresh_data()
            if hasattr(self.parent_form, 'thong_ke_tab'):
                self.parent_form.thong_ke_tab.load_data()
                
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lưu đơn hàng: {str(e)}")
            self.db_session.rollback()

    def export_to_excel(self):
        """Export the order details to Excel"""
        try:
            current_date = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_name = f"truc_in_{current_date}.xlsx"
            
            # Format quy_cach to include unit if not present
            quy_cach = self.quy_cach.get().strip()
            if quy_cach and not any(unit in quy_cach.lower() for unit in ['mm', 'cm', 'm']):
                quy_cach = f"{quy_cach}mm"
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Trục In"
            
            # Prepare data
            data = {
                'Ngày': current_date,
                'Tên Hàng': self.ten_hang_entry.get(),
                'Tên Khách Hàng': self.ten_khach_hang_entry.get(),
                'Ngày dự kiến': self.ngay_du_kien.get_date().strftime('%d-%m-%Y'),
                'Quy cách': quy_cach,
                'Số lượng': self.so_luong.get(),
                'Màu sắc': self.mau_sac.get(),
                'Màu keo': self.mau_keo.get(),
                'Đơn giá gốc': self.truc_in_don_gia_goc.get(),
                'Thành tiền gốc': self.truc_in_thanh_tien.get(),
                'Đơn giá bán': self.truc_in_don_gia_ban.get(),
                'Thành tiền bán': self.truc_in_thanh_tien_ban.get(),
                'Công nợ khách': self.truc_in_cong_no_khach.get(),
                'CTV': self.truc_in_ctv.get(),
                'Hoa hồng': self.truc_in_hoa_hong.get(),
                'Tiền hoa hồng': self.truc_in_tien_hoa_hong.get(),
                'Lợi nhuận': self.truc_in_loi_nhuan.get(),
                'Tiền ship': self.truc_in_tien_ship.get(),
                'Lợi nhuận ròng': self.truc_in_loi_nhuan_rong.get()
            }
            
            # Write headers and data
            for col, (header, value) in enumerate(data.items(), start=1):
                ws.cell(row=1, column=col, value=header)
                ws.cell(row=2, column=col, value=value)
            
            # Save file
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=file_name
            )
            
            if file_path:
                wb.save(file_path)
                messagebox.showinfo("Thành công", f"Đã xuất file Excel: {file_path}")
        
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi xuất Excel: {str(e)}")
            raise

    def export_truc_in_email(self):
        try:
            ten_hang = self.ten_hang_entry.get()
            ten_khach_hang = self.ten_khach_hang_entry.get()
            mau_sac = self.mau_sac.get()
            mau_keo = self.mau_keo.get()
            
            # Format quy_cach to include unit if not present
            quy_cach = self.quy_cach.get().strip()
            if quy_cach and not any(unit in quy_cach.lower() for unit in ['mm', 'cm', 'm']):
                quy_cach = f"{quy_cach}mm"
                
            so_luong = self.so_luong.get()

            content = f"""
THÔNG TIN ĐƠN HÀNG TRỤC IN:
--------------------------
Tên hàng: {ten_hang}
Tên khách hàng: {ten_khach_hang}
Màu sắc: {mau_sac}
Màu keo: {mau_keo}
Quy cách: {quy_cach}
Số lượng: {so_luong}
"""

            # Save to temp file
            current_date = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_name = f"truc_in_{current_date}.txt"
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt")],
                initialfile=file_name
            )
            
            if file_path:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(content)
                messagebox.showinfo("Thành công", f"Đã xuất file: {file_path}")
        
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi xuất file: {str(e)}")
            raise

    def xoa_form_truc_in(self):
        """Clear all form fields"""
        try:
            # Clear basic information
            self.ten_hang_entry.delete(0, tk.END)
            self.ten_khach_hang_entry.delete(0, tk.END)
            self.quy_cach.delete(0, tk.END)
            self.so_luong.delete(0, tk.END)
            self.mau_sac.delete(0, tk.END)
            self.mau_keo.delete(0, tk.END)
            
            # Clear prices
            self.truc_in_don_gia_goc.delete(0, tk.END)
            self.truc_in_thanh_tien.delete(0, tk.END)
            self.truc_in_don_gia_ban.delete(0, tk.END)
            self.truc_in_thanh_tien_ban.delete(0, tk.END)
            self.truc_in_cong_no_khach.delete(0, tk.END)
            
            # Clear CTV and commission
            self.truc_in_ctv.delete(0, tk.END)
            self.truc_in_hoa_hong.delete(0, tk.END)
            self.truc_in_tien_hoa_hong.delete(0, tk.END)
            
            # Clear additional info
            self.truc_in_loi_nhuan.delete(0, tk.END)
            self.truc_in_tien_ship.delete(0, tk.END)
            self.truc_in_loi_nhuan_rong.delete(0, tk.END)
            
            # Reset date to today
            self.ngay_du_kien.set_date(datetime.now())
            
            # Set focus to first field
            self.ten_hang_entry.focus_set()
            
            messagebox.showinfo("Thành công", "Đã xóa form")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi xóa form: {str(e)}")
            raise
