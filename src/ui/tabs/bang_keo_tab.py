# bang_keo_tab.py
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from src.ui.tabs.tab_base import TabBase
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from tkcalendar import DateEntry
from src.database.database import BangKeoOrder, OrderAttachment
from src.utils.autocomplete import AutocompleteEntry

class BangKeoTab(TabBase):
    def __init__(self, container, parent_form):
        super().__init__(parent_form)
        self.container = container
        self.parent_form = parent_form
        self.db_session = parent_form.db_session

        # Create main frame with padding
        main_frame = ttk.Frame(self.container, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Configure grid weights for auto-resizing
        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_columnconfigure(1, weight=1)
        main_frame.grid_columnconfigure(2, weight=1)
        main_frame.grid_columnconfigure(3, weight=1)
        
        for i in range(4):  # 4 rows for different sections
            main_frame.grid_rowconfigure(i, weight=1)

        # Build the UI components
        self.build_ui(main_frame)
        self.bind_events()
        self.bind_currency_format()
        self.bind_shortcuts()
        self.load_suggestions()

        # Checkbox for status
        self.da_giao = tk.BooleanVar(value=False)
        self.da_tat_toan_var = tk.BooleanVar(value=False)

    def build_ui(self, main_frame):
        # Configure grid columns with padding
        main_frame.configure(padding="20")

        # Title with better styling
        title_label = ttk.Label(main_frame, text="Băng Keo", style='Header.TLabel')
        title_label.grid(row=0, column=0, columnspan=4, pady=(0, 20), sticky='nsew')

        # Basic Information Frame
        basic_info_frame = ttk.LabelFrame(main_frame, text="Thông tin cơ bản", padding=15)
        basic_info_frame.grid(row=1, column=0, columnspan=4, sticky='nsew', padx=5, pady=5)
        
        # Configure grid for basic info frame
        basic_info_frame.columnconfigure(0, weight=1, minsize=130)
        basic_info_frame.columnconfigure(1, weight=3)
        basic_info_frame.columnconfigure(2, weight=1, minsize=130)
        basic_info_frame.columnconfigure(3, weight=3)
        for i in range(3):
            basic_info_frame.rowconfigure(i, weight=1)

        # Row 0: Tên hàng và Ngày dự kiến
        ttk.Label(basic_info_frame, text="Tên hàng:").grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.ten_hang_entry = AutocompleteEntry(basic_info_frame, callback=self.auto_fill_data)
        self.ten_hang_entry.grid(row=0, column=1, sticky='ew', padx=5, pady=5)

        ttk.Label(basic_info_frame, text="Ngày dự kiến:").grid(row=0, column=2, sticky='e', padx=5, pady=5)
        self.ngay_du_kien = DateEntry(basic_info_frame, width=20, foreground='white', borderwidth=2)
        self.ngay_du_kien.grid(row=0, column=3, sticky='ew', padx=5, pady=5)

        # Row 1: Tên khách hàng và Quy cách
        ttk.Label(basic_info_frame, text="Tên khách hàng:").grid(row=1, column=0, sticky='e', padx=5, pady=5)
        self.ten_khach_hang_entry = ttk.Entry(basic_info_frame)
        self.ten_khach_hang_entry.grid(row=1, column=1, sticky='ew', padx=5, pady=5)

        ttk.Label(basic_info_frame, text="Quy cách:").grid(row=1, column=2, sticky='e', padx=5, pady=5)
        self.quy_cach = ttk.Entry(basic_info_frame)
        self.quy_cach.grid(row=1, column=3, sticky='ew', padx=5, pady=5)

        # Row 2: Số lượng và Màu sắc
        ttk.Label(basic_info_frame, text="Số lượng:").grid(row=2, column=0, sticky='e', padx=5, pady=5)
        self.so_luong = ttk.Entry(basic_info_frame)
        self.so_luong.grid(row=2, column=1, sticky='ew', padx=5, pady=5)

        ttk.Label(basic_info_frame, text="Màu sắc:").grid(row=2, column=2, sticky='e', padx=5, pady=5)
        self.mau_sac = ttk.Entry(basic_info_frame)
        self.mau_sac.grid(row=2, column=3, sticky='ew', padx=5, pady=5)

        # Price Frame
        price_frame = ttk.LabelFrame(main_frame, text="Giá và Chi phí", padding=10)
        price_frame.grid(row=2, column=0, columnspan=4, sticky='nsew', padx=5, pady=5)
        
        # Configure grid for price frame
        price_frame.columnconfigure(0, weight=1, minsize=130)
        price_frame.columnconfigure(1, weight=3)
        price_frame.columnconfigure(2, weight=1, minsize=130)
        price_frame.columnconfigure(3, weight=3)
        for i in range(5):
            price_frame.rowconfigure(i, weight=1)

        # Đơn giá gốc
        ttk.Label(price_frame, text="Đơn giá gốc:").grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.don_gia_goc = ttk.Entry(price_frame, width=20)
        self.don_gia_goc.grid(row=0, column=1, sticky='ew', padx=5, pady=5)

        # Thành tiền
        ttk.Label(price_frame, text="Thành tiền:").grid(row=1, column=0, sticky='e', padx=5, pady=5)
        self.thanh_tien = ttk.Entry(price_frame, width=20, state='readonly')
        self.thanh_tien.grid(row=1, column=1, sticky='ew', padx=5, pady=5)

        # Đơn giá bán
        ttk.Label(price_frame, text="Đơn giá (bán):").grid(row=0, column=2, sticky='e', padx=5, pady=5)
        self.don_gia_ban = ttk.Entry(price_frame, width=20)
        self.don_gia_ban.grid(row=0, column=3, sticky='ew', padx=5, pady=5)

        # Thành tiền bán
        ttk.Label(price_frame, text="Thành tiền (bán):").grid(row=1, column=2, sticky='e', padx=5, pady=5)
        self.thanh_tien_ban = ttk.Entry(price_frame, width=20, state='readonly')
        self.thanh_tien_ban.grid(row=1, column=3, sticky='ew', padx=5, pady=5)

        # Công nợ khách
        ttk.Label(price_frame, text="Công nợ khách:").grid(row=2, column=0, sticky='e', padx=5, pady=5)
        self.cong_no_khach = ttk.Entry(price_frame, width=20, state='readonly')
        self.cong_no_khach.grid(row=2, column=1, sticky='ew', padx=5, pady=5)

        # CTV và Hoa hồng
        ttk.Label(price_frame, text="CTV:").grid(row=3, column=0, sticky='e', padx=5, pady=5)
        self.ctv = ttk.Entry(price_frame, width=20)
        self.ctv.grid(row=3, column=1, sticky='ew', padx=5, pady=5)
        ttk.Label(price_frame, text="Hoa Hồng (%):").grid(row=3, column=2, sticky='e', padx=5, pady=5)
        self.hoa_hong = ttk.Entry(price_frame, width=20)
        self.hoa_hong.grid(row=3, column=3, sticky='ew', padx=5, pady=5)

        # Tiền hoa hồng
        ttk.Label(price_frame, text="Tiền hoa hồng:").grid(row=4, column=0, sticky='e', padx=5, pady=5)
        self.tien_hoa_hong = ttk.Entry(price_frame, width=20, state='readonly')
        self.tien_hoa_hong.grid(row=4, column=1, sticky='ew', padx=5, pady=5)

        # Lợi nhuận
        ttk.Label(price_frame, text="Lợi nhuận:").grid(row=5, column=0, sticky='e', padx=5, pady=5)
        self.loi_nhuan = ttk.Entry(price_frame, width=20, state='readonly')
        self.loi_nhuan.grid(row=5, column=1, sticky='ew', padx=5, pady=5)

        # Tiền ship
        ttk.Label(price_frame, text="Tiền ship:").grid(row=6, column=0, sticky='e', padx=5, pady=5)
        self.tien_ship = ttk.Entry(price_frame, width=20)
        self.tien_ship.grid(row=6, column=1, sticky='ew', padx=5, pady=5)

        # Lợi nhuận ròng
        ttk.Label(price_frame, text="Lợi nhuận ròng:").grid(row=6, column=2, sticky='e', padx=5, pady=5)
        self.loi_nhuan_rong = ttk.Entry(price_frame, width=20, state='readonly')
        self.loi_nhuan_rong.grid(row=6, column=3, sticky='ew', padx=5, pady=5)

        # Buttons Frame
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=3, column=0, columnspan=4, pady=20, sticky='e')  # Align to the right

        # Define a consistent style for all buttons
        style = ttk.Style()
        style.configure('CustomButton.TButton',
                        font=('Segoe UI', 10),
                        padding=6)

        # Create buttons with the custom style
        btn_tinh_toan = ttk.Button(button_frame, text="Tính toán", command=self.tinh_toan_bang_keo, style='CustomButton.TButton', width=12)
        btn_luu = ttk.Button(button_frame, text="Lưu", command=self.luu_bang_keo, style='CustomButton.TButton', width=12)
        btn_xuat_excel = ttk.Button(button_frame, text="Xuất Excel", command=self.export_to_excel, style='CustomButton.TButton', width=12)
        btn_xuat_email = ttk.Button(button_frame, text="Gửi Email", command=self.export_bang_keo_email, style='CustomButton.TButton', width=12)
        btn_xoa = ttk.Button(button_frame, text="Xóa", command=self.xoa_form_bang_keo, style='CustomButton.TButton', width=12)

        # Pack buttons to the right with consistent padding
        btn_xoa.pack(side='right', padx=5)
        btn_xuat_email.pack(side='right', padx=5)
        btn_xuat_excel.pack(side='right', padx=5)
        btn_luu.pack(side='right', padx=5)
        btn_tinh_toan.pack(side='right', padx=5)

        # Set focus to the first entry
        self.ten_hang_entry.focus_set()

    def bind_events(self):
        entries_to_bind = [
            self.so_luong,
            self.don_gia_ban,
            self.don_gia_goc,
            self.hoa_hong,
            self.tien_ship,
        ]
        for entry in entries_to_bind:
            entry.bind('<KeyRelease>', self.auto_calculate)

        # Register validation command
        vcmd = (self.root.register(self.is_valid_float), '%P')
        for entry in entries_to_bind:
            entry.config(validate='key', validatecommand=vcmd)

    def auto_calculate(self, event):
        self.tinh_toan_bang_keo()

    def bind_currency_format(self):
        currency_fields = [
            self.don_gia_goc,
            self.don_gia_ban,
        ]
        for field in currency_fields:
            field.bind('<FocusOut>', self.format_currency_input)

    def bind_shortcuts(self):
        self.root.bind('<Control-s>', lambda event: self.luu_bang_keo())
        self.root.bind('<Control-t>', lambda event: self.tinh_toan_bang_keo())
        self.root.bind('<Control-e>', lambda event: self.export_to_excel())
        self.root.bind('<Control-q>', lambda event: self.root.quit())

    def tinh_toan_bang_keo(self):
        try:
            # Get values
            so_luong = self.validate_float_input(self.so_luong.get())
            don_gia_ban = self.validate_float_input(self.don_gia_ban.get())
            don_gia_goc = self.validate_float_input(self.don_gia_goc.get())
            hoa_hong = self.validate_float_input(self.hoa_hong.get()) / 100
            tien_ship = self.validate_float_input(self.tien_ship.get())

            # Calculate values
            thanh_tien = don_gia_goc * so_luong
            thanh_tien_ban = don_gia_ban * so_luong
            loi_nhuan = thanh_tien_ban - thanh_tien
            tien_hoa_hong = loi_nhuan * hoa_hong
            cong_no_khach = thanh_tien_ban
            loi_nhuan_rong = loi_nhuan - tien_hoa_hong - tien_ship

            # Update readonly fields
            self.update_readonly_field(self.thanh_tien, thanh_tien)
            self.update_readonly_field(self.thanh_tien_ban, thanh_tien_ban)
            self.update_readonly_field(self.cong_no_khach, cong_no_khach)
            self.update_readonly_field(self.tien_hoa_hong, tien_hoa_hong)
            self.update_readonly_field(self.loi_nhuan, loi_nhuan)
            self.update_readonly_field(self.loi_nhuan_rong, loi_nhuan_rong)

            # self.update_status("Tính toán Băng Keo thành công")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi tính toán Băng Keo: {str(e)}")
            self.update_status("Lỗi khi tính toán Băng Keo")

    def luu_bang_keo(self):
        """Save the Bang Keo order to database"""
        try:
            # Kiểm tra các trường bắt buộc
            required_fields = {
                'Tên hàng': self.ten_hang_entry.get(),
                'Tên khách hàng': self.ten_khach_hang_entry.get(),
                'Số lượng': self.so_luong.get(),
                'Đơn giá bán': self.don_gia_ban.get()
            }
            
            # Validate required fields
            for field_name, value in required_fields.items():
                if not value:
                    messagebox.showerror("Lỗi", f"Vui lòng nhập {field_name}")
                    return
            
            # Create data dictionary
            data = {
                'thoi_gian': datetime.now(),
                'ten_hang': self.ten_hang_entry.get(),
                'ten_khach_hang': self.ten_khach_hang_entry.get(),
                'ngay_du_kien': self.ngay_du_kien.get_date(),
                'quy_cach': self.quy_cach.get(),
                'so_luong': self.parse_float(self.so_luong.get()),
                'mau_sac': self.mau_sac.get(),
                'don_gia_goc': self.parse_float(self.don_gia_goc.get()),
                'thanh_tien': self.parse_float(self.thanh_tien.get()),
                'don_gia_ban': self.parse_float(self.don_gia_ban.get()),
                'thanh_tien_ban': self.parse_float(self.thanh_tien_ban.get()),
                'cong_no_khach': self.parse_float(self.cong_no_khach.get()),
                'ctv': self.ctv.get(),
                'hoa_hong': self.parse_float(self.hoa_hong.get()),
                'tien_hoa_hong': self.parse_float(self.tien_hoa_hong.get()),
                'loi_nhuan': self.parse_float(self.loi_nhuan.get()),
                'tien_ship': self.parse_float(self.tien_ship.get()),
                'loi_nhuan_rong': self.parse_float(self.loi_nhuan_rong.get()),
                'da_giao': False,
                'da_tat_toan': False
            }
            
            # Create new BangKeoOrder object
            don_hang = BangKeoOrder(**data)
            
            # Add to database
            self.db_session.add(don_hang)
            self.db_session.commit()
            
            messagebox.showinfo("Thành công", "Đã lưu đơn hàng thành công!")

            # Hỏi người dùng có muốn đính kèm tệp không
            if messagebox.askyesno("Đính kèm", "Bạn có muốn đính kèm tệp vào đơn hàng này?"):
                file_paths = filedialog.askopenfilenames(title="Chọn tệp đính kèm")
                if file_paths:
                    self._save_attachments(order_id=don_hang.id, order_type='bang_keo', file_paths=file_paths)
            
            # Update history tab and statistics tab
            if hasattr(self.parent_form, 'history_tab'):
                self.parent_form.history_tab.refresh_data()
            if hasattr(self.parent_form, 'thong_ke_tab'):
                self.parent_form.thong_ke_tab.load_data()
                
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lưu đơn hàng: {str(e)}")
            self.db_session.rollback()

    def export_to_excel(self):
        """Export the order details to Excel"""
        try:
            # Get the file path
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"bang_keo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not file_path:
                return

            # Create or load workbook
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                ws = wb.active
                next_row = ws.max_row + 1  # Get the next available row
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Băng Keo"
                next_row = 1  # Start with first row for new file
                
                # Write headers if new file
                headers = [
                    'ID', 'Ngày', 'Tên Hàng', 'Tên Khách Hàng', 'Ngày dự kiến',
                    'Quy cách', 'Số lượng', 'Màu sắc', 'Đơn giá gốc',
                    'Thành tiền', 'Đơn giá bán', 'Thành tiền bán',
                    'Công nợ khách', 'CTV', 'Hoa hồng', 'Tiền hoa hồng',
                    'Lợi nhuận', 'Tiền ship', 'Lợi nhuận ròng'
                ]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                next_row = 2  # Start data from second row
            
            # Prepare data for the current order
            current_date = datetime.now().strftime('%m/%d/%Y')
            data = [
                self.ten_hang_entry.get(),  # ID/Tên hàng for now
                current_date,
                self.ten_hang_entry.get(),
                self.ten_khach_hang_entry.get(),
                self.ngay_du_kien.get_date().strftime('%m/%d/%Y'),
                self.quy_cach.get(),
                self.so_luong.get(),
                self.mau_sac.get(),
                self.don_gia_goc.get(),
                self.thanh_tien.get(),
                self.don_gia_ban.get(),
                self.thanh_tien_ban.get(),
                self.cong_no_khach.get(),
                self.ctv.get(),
                self.hoa_hong.get(),
                self.tien_hoa_hong.get(),
                self.loi_nhuan.get(),
                self.tien_ship.get(),
                self.loi_nhuan_rong.get()
            ]
            
            # Write data to the next available row
            for col, value in enumerate(data, 1):
                ws.cell(row=next_row, column=col, value=value)
            
            # Save workbook
            wb.save(file_path)
            messagebox.showinfo("Thành công", f"Đã xuất file Excel: {file_path}")
        
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi xuất Excel: {str(e)}")
            raise

    def _format_number(self, value):
        """Format số: nếu là số nguyên thì không hiển thị .0, chỉ hiển thị phần thập phân nếu có"""
        if value is None or value == '':
            return ''
        try:
            num = float(value)
            if num.is_integer():
                return str(int(num))
            return str(num)
        except (ValueError, TypeError):
            return str(value)
    
    def export_bang_keo_email(self):
        try:
            from src.ui.tabs.history_components.email_dialog import EmailDialog
            
            ten_hang = self.ten_hang_entry.get()
            ten_khach_hang = self.ten_khach_hang_entry.get()
            mau_sac = self.mau_sac.get()
            quy_cach = self._format_number(self.quy_cach.get())
            so_luong = self._format_number(self.so_luong.get())
            
            subject = f"Đơn hàng Băng Keo: {ten_hang}" if ten_hang else "Đơn hàng Băng Keo"
            content = f"""Chào bác,

Bác làm giúp con đơn hàng băng keo bên dưới nhé:

THÔNG TIN ĐƠN HÀNG BĂNG KEO:
--------------------------
Tên hàng: {ten_hang}
Tên khách hàng: {ten_khach_hang}
Màu sắc: {mau_sac}
Quy cách (KG): {quy_cach}
Số lượng: {so_luong}

Cảm ơn bác!
Quế 
"""
            # Mở dialog email (không có order_id vì đơn chưa lưu)
            dlg = EmailDialog(self.root, self.db_session, 'bang_keo', 'temp', subject, content)
            self.root.wait_window(dlg)
        
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi mở dialog email: {str(e)}")
            raise

    def xoa_form_bang_keo(self):
        """Clear all form fields"""
        try:
            # Clear basic information
            self.ten_hang_entry.delete(0, tk.END)
            self.ten_khach_hang_entry.delete(0, tk.END)
            self.quy_cach.delete(0, tk.END)
            self.so_luong.delete(0, tk.END)
            self.mau_sac.delete(0, tk.END)
            
            # Clear prices
            self.don_gia_goc.delete(0, tk.END)
            self.thanh_tien.delete(0, tk.END)
            self.don_gia_ban.delete(0, tk.END)
            self.thanh_tien_ban.delete(0, tk.END)
            self.cong_no_khach.delete(0, tk.END)
            
            # Clear CTV and commission
            self.ctv.delete(0, tk.END)
            self.hoa_hong.delete(0, tk.END)
            self.tien_hoa_hong.delete(0, tk.END)
            
            # Clear additional info
            self.loi_nhuan.delete(0, tk.END)
            self.tien_ship.delete(0, tk.END)
            self.loi_nhuan_rong.delete(0, tk.END)
            
            # Reset date to today
            self.ngay_du_kien.set_date(datetime.now())
            
            # Set focus to first field
            self.ten_hang_entry.focus_set()
            
            messagebox.showinfo("Thành công", "Đã xóa form")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi xóa form: {str(e)}")
            raise 

    def _save_attachments(self, order_id, order_type, file_paths):
        try:
            saved = 0
            for path in file_paths:
                try:
                    with open(path, 'rb') as f:
                        data = f.read()
                    file_name = os.path.basename(path)
                    ext = os.path.splitext(file_name)[1].lower()
                    content_type = {
                        '.pdf': 'application/pdf',
                        '.png': 'image/png',
                        '.jpg': 'image/jpeg',
                        '.jpeg': 'image/jpeg',
                        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        '.csv': 'text/csv',
                        '.txt': 'text/plain'
                    }.get(ext, 'application/octet-stream')

                    att = OrderAttachment(
                        order_type=order_type,
                        order_id=order_id,
                        file_name=file_name,
                        content_type=content_type,
                        file_size=len(data),
                        data=data
                    )
                    self.db_session.add(att)
                    saved += 1
                except Exception as item_err:
                    messagebox.showwarning("Cảnh báo", f"Không thể lưu tệp: {path}\nLý do: {item_err}")

            self.db_session.commit()
            if saved:
                messagebox.showinfo("Thành công", f"Đã lưu {saved} tệp đính kèm lên cơ sở dữ liệu")
        except Exception as e:
            self.db_session.rollback()
            messagebox.showerror("Lỗi", f"Không thể lưu tệp đính kèm: {str(e)}")

    def load_suggestions(self):
        """Load unique item names for autocomplete"""
        try:
            suggestions = self.db_session.query(BangKeoOrder.ten_hang).distinct().all()
            suggestion_list = [s[0] for s in suggestions if s[0]]
            self.ten_hang_entry.set_suggestions(suggestion_list)
        except Exception as e:
            print(f"Error loading suggestions: {e}")

    def auto_fill_data(self, ten_hang):
        """Auto-fill form based on the most recent order of the selected item"""
        if not ten_hang:
            return
            
        try:
            last_order = (
                self.db_session.query(BangKeoOrder)
                .filter(BangKeoOrder.ten_hang == ten_hang)
                .order_by(BangKeoOrder.thoi_gian.desc())
                .first()
            )
            
            if last_order:
                # Update basic info
                self.ten_khach_hang_entry.delete(0, tk.END)
                self.ten_khach_hang_entry.insert(0, last_order.ten_khach_hang or "")
                
                self.quy_cach.delete(0, tk.END)
                self.quy_cach.insert(0, last_order.quy_cach or "")
                
                self.so_luong.delete(0, tk.END)
                self.so_luong.insert(0, self._format_number(last_order.so_luong))
                
                self.mau_sac.delete(0, tk.END)
                self.mau_sac.insert(0, last_order.mau_sac or "")
                
                self.don_gia_goc.delete(0, tk.END)
                self.don_gia_goc.insert(0, self.format_currency(last_order.don_gia_goc or 0))
                
                self.don_gia_ban.delete(0, tk.END)
                self.don_gia_ban.insert(0, self.format_currency(last_order.don_gia_ban or 0))
                
                self.ctv.delete(0, tk.END)
                self.ctv.insert(0, last_order.ctv or "")
                
                self.hoa_hong.delete(0, tk.END)
                self.hoa_hong.insert(0, self._format_number(last_order.hoa_hong))
                
                self.tien_ship.delete(0, tk.END)
                self.tien_ship.insert(0, self.format_currency(last_order.tien_ship or 0))
                
                # Trigger calculation
                self.tinh_toan_bang_keo()
                self.update_status(f"Đã tự động điền thông tin từ đơn hàng cũ cho '{ten_hang}'")
                
        except Exception as e:
            print(f"Error auto-filling data: {e}")
            self.update_status("Lỗi khi tự động điền thông tin")
